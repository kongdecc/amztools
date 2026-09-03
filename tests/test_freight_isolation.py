"""Regression tests: independent browser snapshots, no database or shared files."""
import base64
import importlib.util
import io
import json
import sys
import unittest
import zipfile
from concurrent.futures import ThreadPoolExecutor
from pathlib import Path
from unittest.mock import patch

from openpyxl import Workbook, load_workbook

ROOT = Path(__file__).resolve().parents[1]
spec = importlib.util.spec_from_file_location('freight_function', ROOT / 'api/freight-invoice-python.py')
backend = importlib.util.module_from_spec(spec)
spec.loader.exec_module(backend)


class Browser:
    def __init__(self):
        self.snapshot = ''

    def call(self, path, body=None, method=None):
        payload = body if isinstance(body, bytes) else json.dumps(body).encode() if body is not None else b''
        result = backend.execute({'protocol': backend.PROTOCOL, 'path': path,
                                  'method': method or ('POST' if body is not None else 'GET'),
                                  'body': backend.encode(payload), 'snapshot': self.snapshot})
        if 200 <= result['status'] < 300 and 'snapshot' in result:
            self.snapshot = result['snapshot']
        content = base64.b64decode(result['body'])
        if 'application/json' in result['headers'].get('Content-Type', ''):
            content = json.loads(content)
        return result['status'], content


class IsolationTests(unittest.TestCase):
    def test_products_drafts_fields_and_cold_start(self):
        a, b = Browser(), Browser()
        a.call('/products', {'products': [{'id': 'a', 'sku': 'A-only'}]})
        a.call('/draft', {'items': [{'sku': 'A-only'}], 'shipment': {}})
        a.call('/field-catalog', {'category': 'items', 'label': 'A-only field', 'inputType': 'text'})
        self.assertEqual(b.call('/products')[1]['products'], [])
        self.assertIsNone(b.call('/draft')[1]['draft'])
        self.assertFalse(any(f['label'] == 'A-only field' for f in b.call('/field-catalog')[1]['items']))
        self.assertTrue(any(f['label'] == 'A-only field' for f in a.call('/field-catalog')[1]['items']))
        reopened = Browser()
        reopened.snapshot = a.snapshot
        self.assertEqual(reopened.call('/products')[1]['products'][0]['sku'], 'A-only')
        b.call('/products', {'products': [{'sku': 'B-only'}]})
        a.call('/data/clear-all', {'confirmation': 'CLEAR_ALL_DATA'})
        self.assertEqual(a.call('/products')[1]['products'], [])
        self.assertEqual(b.call('/products')[1]['products'][0]['sku'], 'B-only')

    def test_template_preview_export_history_backup_restore(self):
        a, b = Browser(), Browser()
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = 'Invoice'
        sheet.append(['SKU', 'Product Name', 'Quantity'])
        sheet.append(['', '', ''])
        file = io.BytesIO()
        workbook.save(file)
        workbook.close()
        status, uploaded = a.call('/templates/upload', {'name': 'A template', 'filename': 'test.xlsx', 'data': backend.encode(file.getvalue())})
        self.assertEqual(status, 200, uploaded)
        template_id = uploaded['template']['id']
        mapping = {'sheet': 'Invoice', 'fixed': {}, 'required': {'fixed': [], 'items': []},
                   'items': {'startRow': 2, 'headerRow': 1, 'reservedRows': 1, 'rowMode': 'sku',
                             'columns': {'exportSku': 'A', 'nameEn': 'B', 'quantity': 'C'}}}
        self.assertEqual(a.call(f'/templates/{template_id}/mapping', {'mapping': mapping})[0], 200)
        self.assertEqual(b.call('/templates')[1]['templates'], [])
        self.assertEqual(b.call(f'/templates/{template_id}')[0], 404)
        payload = {'templateId': template_id, 'shipment': {}, 'items': [{'exportSku': 'PRIVATE', 'nameEn': 'A product', 'quantity': 2, 'cartons': 1}]}
        self.assertEqual(a.call('/preview', payload)[0], 200)
        status, exported = a.call('/export', payload)
        self.assertEqual(status, 200, exported)
        result_book = load_workbook(io.BytesIO(exported))
        self.assertEqual(result_book.active['A2'].value, 'PRIVATE')
        result_book.close()
        history = a.call('/history')[1]['history']
        self.assertEqual(len(history), 1)
        self.assertEqual(b.call('/history')[1]['history'], [])
        self.assertEqual(a.call(f"/history/{history[0]['id']}/file")[1], exported)
        backup = a.call('/backup')[1]
        with zipfile.ZipFile(io.BytesIO(backup)) as archive:
            self.assertNotIn('warehouses.json', archive.namelist())
        b.call('/products', {'products': [{'sku': 'B-only'}]})
        a.call('/data/clear-all', {'confirmation': 'CLEAR_ALL_DATA'})
        self.assertEqual(a.call('/restore', backup)[0], 200)
        self.assertEqual(len(a.call('/templates')[1]['templates']), 1)
        self.assertEqual(b.call('/products')[1]['products'][0]['sku'], 'B-only')
        original = a.snapshot
        self.assertEqual(a.call('/restore', b'bad zip')[0], 400)
        self.assertEqual(a.snapshot, original)

    def test_no_snapshot_no_data_and_legacy_rejected(self):
        a = Browser()
        a.call('/products', {'products': [{'sku': 'secret'}]})
        self.assertEqual(Browser().call('/products')[1]['products'], [])
        with self.assertRaises(ValueError):
            backend.execute({'path': '/products'})
        with self.assertRaises(ValueError):
            backend.execute({'protocol': backend.PROTOCOL, 'path': '/warehouses'})

    def test_parallel_isolation_and_temporary_cleanup(self):
        def run(name):
            browser = Browser()
            for _ in range(3):
                browser.call('/products', {'products': [{'sku': name}]})
                self.assertEqual(browser.call('/products')[1]['products'][0]['sku'], name)
        with ThreadPoolExecutor(max_workers=4) as pool:
            list(pool.map(run, ['A', 'B', 'C', 'D']))
        folders = []
        def fail(*args):
            folders.append(backend.engine.DATA_DIR)
            raise ValueError('test failure')
        with patch.object(backend.engine, 'generate_workbook', fail):
            self.assertEqual(Browser().call('/preview', {'items': []})[0], 400)
        self.assertTrue(folders)
        self.assertTrue(all(not folder.exists() for folder in folders))


if __name__ == '__main__':
    unittest.main()
