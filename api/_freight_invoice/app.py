from __future__ import annotations

import base64
import copy
import io
import json
import mimetypes
import os
import re
import shutil
import subprocess
import sys
import tempfile
import threading
import uuid
import webbrowser
import zipfile
from datetime import datetime
from http import HTTPStatus
from http.server import SimpleHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path, PurePosixPath
from urllib.parse import parse_qs, unquote, urlparse
from urllib.request import ProxyHandler, build_opener

from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.formula.translate import Translator
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.utils.units import pixels_to_EMU, points_to_pixels


APP_VERSION = "3.0.0-beta"
APP_EDITION = "通用服务器版"
BUNDLE_ROOT = Path(getattr(sys, "_MEIPASS", Path(__file__).resolve().parent))
APP_ROOT = Path(sys.executable).resolve().parent if getattr(sys, "frozen", False) else BUNDLE_ROOT
VENDOR_DIR = BUNDLE_ROOT / "vendor"
if VENDOR_DIR.is_dir() and str(VENDOR_DIR) not in sys.path:
    sys.path.insert(0, str(VENDOR_DIR))
STATIC_DIR = BUNDLE_ROOT / "static"
TEMPLATE_DIR = BUNDLE_ROOT / "templates"
EXPORT_DIR = APP_ROOT / "exports"
DATA_DIR = Path(os.environ.get("FREIGHT_INVOICE_DATA_DIR", str(APP_ROOT / "data"))).expanduser().resolve()
PRODUCTS_FILE = DATA_DIR / "products.json"
WAREHOUSES_FILE = DATA_DIR / "warehouses.json"
WAREHOUSE_DEFAULTS_FILE = TEMPLATE_DIR / "warehouses.json"
HISTORY_FILE = DATA_DIR / "export_history.json"
HISTORY_FILES_DIR = DATA_DIR / "history_files"
DRAFT_FILE = DATA_DIR / "invoice_draft.json"
CUSTOM_TEMPLATES_FILE = DATA_DIR / "custom_templates.json"
CUSTOM_TEMPLATES_DIR = DATA_DIR / "custom_templates"
TEMPLATE_SETTINGS_FILE = DATA_DIR / "template_settings.json"
HISTORY_LOCK = threading.Lock()
DRAFT_LOCK = threading.Lock()
TEMPLATE_LOCK = threading.Lock()
TEMPLATE_REQUIRED_CACHE = {}
WAREHOUSE_LOCK = threading.Lock()
BACKUP_ROOT_FILES = {
    "products.json", "export_history.json", "invoice_draft.json",
    "custom_templates.json", "template_settings.json",
}
BACKUP_IGNORED_FILES = {"warehouses.json"}
BACKUP_ROOT_DIRS = {"custom_templates", "history_files"}
BACKUP_MAX_BYTES = 1024 * 1024 * 1024

TEMPLATES = {
    "ruichi": {
        "name": "锐驰 FBA 专线",
        "description": "收件信息、FBA/Reference ID、商品申报明细和合计",
        "file": TEMPLATE_DIR / "锐驰空白模板.xlsx",
    },
    "yuntuo": {
        "name": "广州云拓 B2B 单票",
        "description": "渠道、收件资料、VAT/EORI、逐箱商品申报明细",
        "file": TEMPLATE_DIR / "云拓-b2b单票导入模板.xlsx",
    },
}

FIXED_FIELD_DEFS = [
    ("receiverCompany", "收件公司"), ("receiverName", "收件人"), ("address1", "地址一"),
    ("address2", "地址二"), ("address3", "地址三"), ("fullAddress", "完整收件地址"), ("city", "城市"),
    ("state", "州/省"), ("postalCode", "邮编"), ("countryCode", "国家代码"),
    ("phone", "电话"), ("email", "邮箱"), ("fbaNumber", "FBA/客户订单号"),
    ("amazonReferenceId", "Amazon Reference ID"), ("service", "服务渠道"),
    ("warehouseCode", "地址库编码"), ("hasBattery", "整票带电"), ("hasMagnet", "整票带磁"),
    ("hasLiquid", "整票含液体"), ("hasPowder", "整票含粉末"),
    ("customsMode", "报关方式"), ("clearanceMode", "清关方式"), ("taxMode", "交税方式"),
    ("incoterm", "交货条款"), ("vatCountry", "VAT 注册国家"), ("eori", "EORI"),
    ("vatNumber", "VAT 号"), ("vatName", "VAT 公司英文名"), ("vatAddress", "VAT 注册地址"),
    ("currency", "币种"), ("totalCartons", "总箱数"), ("totalQuantity", "总数量"),
    ("totalGrossWeight", "总毛重"), ("totalNetWeight", "总净重"), ("totalCbm", "总体积"),
    ("totalValue", "申报总金额"), ("notes", "备注"),
]

ITEM_FIELD_DEFS = [
    ("boxNumber", "货箱编号/单号"), ("exportSku", "导出 SKU"), ("nameZh", "中文品名"),
    ("nameEn", "英文品名"), ("quantity", "每箱数量"), ("totalQuantity", "总数量"),
    ("unitPrice", "申报单价"), ("totalAmount", "申报总价"), ("cartons", "箱数"),
    ("grossWeight", "毛重"), ("netWeight", "净重"), ("productWeightText", "产品重量"),
    ("cartonLength", "箱长"), ("cartonWidth", "箱宽"), ("cartonHeight", "箱高"),
    ("cbm", "体积 CBM"), ("hsCode", "海关编码"), ("poNumber", "PO Number"),
    ("brand", "品牌"), ("brandType", "品牌类型"), ("model", "型号"),
    ("material", "材质（中英文）"), ("purpose", "用途（中英文）"), ("image", "产品图片"),
    ("hasMagnet", "是否带磁"), ("hasBattery", "是否带电"), ("batteryInfo", "电池信息"),
    ("appointmentWindow", "预约配送时段"), ("salePrice", "销售价格"), ("saleUrl", "销售链接"),
    ("asin", "产品 ASIN"), ("fnsku", "产品 FNSKU"),
]

FIELD_ALIASES = {
    "receiverCompany": ["收件人公司", "收件公司", "收货公司", "company"],
    "receiverName": ["收件人姓名", "收货人姓名", "收件人", "收货人", "recipient", "consignee"],
    "address1": ["收件人地址一", "收件地址一", "收货地址一", "address1", "address line 1"],
    "address2": ["收件人地址二", "收件地址二", "收货地址二", "address2", "address line 2"],
    "address3": ["收件人地址三", "收件地址三", "收货地址三", "address3", "address line 3"],
    "fullAddress": ["完整收件地址", "完整收货地址", "收件地址", "收货地址", "address"],
    "city": ["收件人城市", "收件城市", "城市", "city"],
    "state": ["收件人省份/州", "收件省份/州", "州代码", "省份", "州/省", "state", "province"],
    "postalCode": ["收件人邮编", "收件邮编", "邮编", "邮政编码", "zip", "postal"],
    "countryCode": ["国家代码", "国家二字码", "country code"],
    "phone": ["收件人电话", "收件电话", "电话", "手机", "phone", "tel"],
    "email": ["收件人邮箱", "收件邮箱", "邮箱", "email"],
    "fbaNumber": ["fba#", "fba号", "fba no", "客户订单号", "客户单号", "shipment id"],
    "amazonReferenceId": ["amazon reference id", "reference id", "reference"],
    "service": ["服务-渠道代码", "服务渠道", "服务", "渠道代码", "渠道", "service"],
    "warehouseCode": ["地址库编码", "仓库地址编码", "warehouse code"],
    "currency": ["申报币种", "币种", "currency"],
    "totalCartons": ["总箱数", "箱数", "total cartons", "ctns"],
    "customsMode": ["报关方式", "customs mode"], "clearanceMode": ["清关方式", "clearance mode"],
    "taxMode": ["交税方式", "tax mode"], "incoterm": ["交货条款", "贸易条款", "incoterm"],
    "vatCountry": ["vat注册国家", "vat country"], "eori": ["eori"],
    "vatNumber": ["vat号", "vat number"], "vatName": ["vat公司英文名", "vat注册名"],
    "vatAddress": ["vat注册地址", "vat address"],
    "hasLiquid": ["液体", "含液体"], "hasPowder": ["粉末", "含粉末"],
    "notes": ["备注", "remark", "notes"],
    "boxNumber": ["货箱编号", "箱号", "单号", "box number", "carton no"],
    "exportSku": ["产品sku", "sku", "货号"], "nameZh": ["中文品名", "产品中文品名"],
    "nameEn": ["英文品名", "产品英文品名", "product name"],
    "quantity": ["每箱数量", "申报数量", "数量", "qty", "pcs"],
    "unitPrice": ["申报单价", "单价", "unit price"], "totalAmount": ["总价", "总金额", "amount"],
    "cartons": ["箱数", "ctns", "cartons"], "grossWeight": ["货箱重量", "箱重", "毛重", "gross weight", "g.w"],
    "netWeight": ["净重", "net weight", "n.w"], "productWeightText": ["产品重量", "单个产品重量"],
    "cartonLength": ["箱长", "货箱长度", "length"], "cartonWidth": ["箱宽", "货箱宽度", "width"],
    "cartonHeight": ["箱高", "货箱高度", "height"], "cbm": ["cbm", "体积"],
    "hsCode": ["海关编码", "hs code", "hscode"], "poNumber": ["po number", "po#"],
    "brand": ["产品品牌", "品牌", "brand"], "brandType": ["品牌类型", "品牌属性", "brand type"],
    "model": ["型号", "model"], "material": ["材质", "material"],
    "purpose": ["用途", "purpose"], "image": ["产品图片", "图片", "image", "photo"],
    "hasMagnet": ["带磁", "磁性"], "hasBattery": ["带电", "电池"],
    "appointmentWindow": ["预约配送时段", "预约时段", "delivery window", "appointment"],
    "salePrice": ["销售价格", "售价"], "saleUrl": ["销售链接", "链接", "url"],
    "asin": ["产品asin", "asin"], "fnsku": ["产品fnsku", "fnsku"],
}


def safe_text(value, default="") -> str:
    if value is None:
        return default
    return str(value).strip()


def read_products() -> list:
    try:
        data = json.loads(PRODUCTS_FILE.read_text(encoding="utf-8"))
        return data if isinstance(data, list) else []
    except (FileNotFoundError, json.JSONDecodeError, OSError):
        return []


def read_warehouses() -> list:
    for source in (WAREHOUSES_FILE, WAREHOUSE_DEFAULTS_FILE):
        try:
            data = json.loads(source.read_text(encoding="utf-8"))
            if isinstance(data, list):
                warehouses = []
                for raw in data:
                    if not isinstance(raw, dict):
                        continue
                    record = dict(raw)
                    if not re.fullmatch(r"wh-[0-9a-f]{12}", safe_text(record.get("id"))):
                        identity = "|".join(safe_text(record.get(key)) for key in ("code", "countryCode", "address", "city", "state", "postalCode"))
                        record["id"] = f"wh-{uuid.uuid5(uuid.NAMESPACE_URL, identity).hex[:12]}"
                    warehouses.append(record)
                return warehouses
        except (FileNotFoundError, json.JSONDecodeError, OSError):
            continue
    return []


def write_warehouses(warehouses: list) -> None:
    DATA_DIR.mkdir(exist_ok=True)
    temporary = WAREHOUSES_FILE.with_suffix(".json.tmp")
    temporary.write_text(json.dumps(warehouses, ensure_ascii=False, indent=2), encoding="utf-8")
    temporary.replace(WAREHOUSES_FILE)


def upsert_warehouse(payload: dict) -> dict:
    code = safe_text(payload.get("code")).upper()
    if not code or len(code) > 30:
        raise ValueError("仓库代码不能为空且不能超过 30 个字符")
    warehouse_id = safe_text(payload.get("id"))
    record = {
        "id": warehouse_id if re.fullmatch(r"wh-[0-9a-f]{12}", warehouse_id) else f"wh-{uuid.uuid4().hex[:12]}",
        "code": code,
        "country": safe_text(payload.get("country")),
        "countryCode": safe_text(payload.get("countryCode")).upper(),
        "region": safe_text(payload.get("region")),
        "address": safe_text(payload.get("address")),
        "city": safe_text(payload.get("city")),
        "state": safe_text(payload.get("state")).upper(),
        "postalCode": safe_text(payload.get("postalCode")),
    }
    with WAREHOUSE_LOCK:
        warehouses = read_warehouses()
        index = next((index for index, value in enumerate(warehouses) if safe_text(value.get("id")) == record["id"]), -1)
        duplicate = next((value for value in warehouses if safe_text(value.get("code")).casefold() == code.casefold() and safe_text(value.get("id")) != record["id"] and safe_text(value.get("address")).casefold() == record["address"].casefold()), None)
        if duplicate:
            raise ValueError("相同代码和地址的仓库已经存在")
        if index >= 0:
            warehouses[index] = record
        else:
            warehouses.append(record)
        write_warehouses(warehouses)
    return record


def delete_warehouse(warehouse_id: str) -> bool:
    if not re.fullmatch(r"wh-[0-9a-f]{12}", warehouse_id):
        return False
    with WAREHOUSE_LOCK:
        warehouses = read_warehouses()
        kept = [record for record in warehouses if safe_text(record.get("id")) != warehouse_id]
        if len(kept) == len(warehouses):
            return False
        write_warehouses(kept)
    return True


def search_warehouses(query: str, limit: int = 30) -> tuple[list, int]:
    words = [part.casefold() for part in safe_text(query).split() if part]
    ranked = []
    for warehouse in read_warehouses():
        code = safe_text(warehouse.get("code"))
        values = [
            code,
            warehouse.get("address"),
            warehouse.get("city"),
            warehouse.get("state"),
            warehouse.get("postalCode"),
            warehouse.get("country"),
            warehouse.get("countryCode"),
            warehouse.get("region"),
        ]
        haystack = " ".join(safe_text(value) for value in values).casefold()
        if words and not all(word in haystack for word in words):
            continue
        normalized_code = re.sub(r"\s*\(.*?\)\s*$", "", code).casefold()
        normalized_query = safe_text(query).casefold()
        rank = 2
        if normalized_query and normalized_code == normalized_query:
            rank = 0
        elif normalized_query and normalized_code.startswith(normalized_query):
            rank = 1
        ranked.append((rank, code.casefold(), warehouse))
    ranked.sort(key=lambda item: (item[0], item[1]))
    matches = [item[2] for item in ranked]
    return matches[:max(1, min(limit, 100))], len(matches)


def write_products(products: list) -> None:
    if not isinstance(products, list):
        raise ValueError("产品资料格式不正确")
    DATA_DIR.mkdir(exist_ok=True)
    temporary = PRODUCTS_FILE.with_suffix(".json.tmp")
    temporary.write_text(json.dumps(products, ensure_ascii=False, indent=2), encoding="utf-8")
    temporary.replace(PRODUCTS_FILE)


def delete_products(product_ids: list) -> int:
    selected = {safe_text(value) for value in product_ids if safe_text(value)}
    if not selected:
        return 0
    products = read_products()
    kept = [product for product in products if safe_text(product.get("id")) not in selected]
    write_products(kept)
    return len(products) - len(kept)


def clear_all_user_data() -> dict:
    """Remove every user-owned record and uploaded/generated file."""
    counts = {
        "products": len(read_products()),
        "templates": len(read_custom_templates()),
        "history": len(read_export_history()),
        "hasDraft": read_invoice_draft() is not None,
    }
    with HISTORY_LOCK, DRAFT_LOCK, TEMPLATE_LOCK:
        for path in (
            PRODUCTS_FILE, HISTORY_FILE, DRAFT_FILE,
            CUSTOM_TEMPLATES_FILE, TEMPLATE_SETTINGS_FILE,
        ):
            try:
                path.unlink(missing_ok=True)
            except OSError:
                pass
        for directory in (CUSTOM_TEMPLATES_DIR, HISTORY_FILES_DIR, EXPORT_DIR):
            if directory.is_dir():
                shutil.rmtree(directory, ignore_errors=True)
    return counts


def read_custom_templates() -> list:
    try:
        data = json.loads(CUSTOM_TEMPLATES_FILE.read_text(encoding="utf-8"))
        return data if isinstance(data, list) else []
    except (FileNotFoundError, json.JSONDecodeError, OSError):
        return []


def write_custom_templates(records: list) -> None:
    DATA_DIR.mkdir(exist_ok=True)
    temporary = CUSTOM_TEMPLATES_FILE.with_suffix(".json.tmp")
    temporary.write_text(json.dumps(records, ensure_ascii=False, indent=2), encoding="utf-8")
    temporary.replace(CUSTOM_TEMPLATES_FILE)


def read_template_settings() -> dict:
    try:
        data = json.loads(TEMPLATE_SETTINGS_FILE.read_text(encoding="utf-8"))
        return data if isinstance(data, dict) else {}
    except (FileNotFoundError, json.JSONDecodeError, OSError):
        return {}


def write_template_settings(settings: dict) -> None:
    DATA_DIR.mkdir(exist_ok=True)
    temporary = TEMPLATE_SETTINGS_FILE.with_suffix(".json.tmp")
    temporary.write_text(json.dumps(settings, ensure_ascii=False, indent=2), encoding="utf-8")
    temporary.replace(TEMPLATE_SETTINGS_FILE)


FIELD_INPUT_TYPES = {"text", "number", "textarea"}


def default_field_input_type(category: str, key: str) -> str:
    if key in {"fullAddress", "address1", "vatAddress", "notes", "purpose", "batteryInfo"}:
        return "textarea"
    numeric_item_fields = {
        "quantity", "totalQuantity", "unitPrice", "totalAmount", "cartons", "grossWeight",
        "netWeight", "cartonLength", "cartonWidth", "cartonHeight", "cbm", "salePrice",
    }
    numeric_fixed_fields = {
        "totalCartons", "totalQuantity", "totalGrossWeight", "totalNetWeight", "totalCbm", "totalValue",
    }
    return "number" if key in (numeric_item_fields if category == "items" else numeric_fixed_fields) else "text"


def field_catalog(include_disabled: bool = False) -> dict[str, list[dict]]:
    settings = read_template_settings().get("fieldCatalog", {})
    overrides = settings.get("overrides", {}) if isinstance(settings, dict) else {}
    custom = settings.get("custom", []) if isinstance(settings, dict) else []
    result = {"fixed": [], "items": []}
    for category, definitions in (("fixed", FIXED_FIELD_DEFS), ("items", ITEM_FIELD_DEFS)):
        for key, default_label in definitions:
            override = overrides.get(f"{category}:{key}", {})
            enabled = override.get("enabled", True) is not False
            aliases = override.get("aliases") if isinstance(override.get("aliases"), list) else FIELD_ALIASES.get(key, [])
            field = {
                "key": key,
                "label": safe_text(override.get("label"), default_label),
                "aliases": list(dict.fromkeys([safe_text(value) for value in aliases if safe_text(value)])),
                "inputType": safe_text(override.get("inputType"), default_field_input_type(category, key)),
                "builtIn": True,
                "enabled": enabled,
            }
            if include_disabled or enabled:
                result[category].append(field)
    for raw in custom if isinstance(custom, list) else []:
        if not isinstance(raw, dict) or raw.get("category") not in result:
            continue
        enabled = raw.get("enabled", True) is not False
        field = {
            "key": safe_text(raw.get("key")),
            "label": safe_text(raw.get("label")),
            "aliases": list(dict.fromkeys([safe_text(value) for value in raw.get("aliases", []) if safe_text(value)])),
            "inputType": safe_text(raw.get("inputType"), "text"),
            "builtIn": False,
            "enabled": enabled,
        }
        if re.fullmatch(r"custom(?:Fixed|Item)_[0-9a-f]{10}", field["key"]) and field["label"] and (include_disabled or enabled):
            result[raw["category"]].append(field)
    return result


def normalize_field_aliases(value, label: str) -> list[str]:
    if isinstance(value, str):
        values = re.split(r"[,，\n]+", value)
    elif isinstance(value, list):
        values = value
    else:
        values = []
    aliases = []
    for candidate in [label, *values]:
        alias = safe_text(candidate)
        if alias and len(alias) <= 100 and alias not in aliases:
            aliases.append(alias)
    return aliases[:30]


def upsert_field_definition(payload: dict) -> dict:
    category = safe_text(payload.get("category"))
    if category not in ("fixed", "items"):
        raise ValueError("字段类别无效")
    label = safe_text(payload.get("label"))
    if not label or len(label) > 80:
        raise ValueError("字段名称不能为空且不能超过 80 个字符")
    input_type = safe_text(payload.get("inputType"), "text")
    if input_type not in FIELD_INPUT_TYPES:
        raise ValueError("字段输入类型无效")
    aliases = normalize_field_aliases(payload.get("aliases"), label)
    key = safe_text(payload.get("key"))
    builtin_keys = {item[0] for item in (FIXED_FIELD_DEFS if category == "fixed" else ITEM_FIELD_DEFS)}
    with TEMPLATE_LOCK:
        settings = read_template_settings()
        catalog_settings = settings.setdefault("fieldCatalog", {})
        overrides = catalog_settings.setdefault("overrides", {})
        custom = catalog_settings.setdefault("custom", [])
        if key in builtin_keys:
            overrides[f"{category}:{key}"] = {
                "label": label, "aliases": aliases, "inputType": input_type, "enabled": True,
            }
        else:
            existing = next((field for field in custom if field.get("category") == category and field.get("key") == key), None)
            if existing is None:
                key = f"custom{'Fixed' if category == 'fixed' else 'Item'}_{uuid.uuid4().hex[:10]}"
                existing = {"category": category, "key": key}
                custom.append(existing)
            existing.update({"label": label, "aliases": aliases, "inputType": input_type, "enabled": True})
        write_template_settings(settings)
    return next(field for field in field_catalog(include_disabled=True)[category] if field["key"] == key)


def delete_field_definition(category: str, key: str) -> bool:
    if category not in ("fixed", "items"):
        return False
    builtin_keys = {item[0] for item in (FIXED_FIELD_DEFS if category == "fixed" else ITEM_FIELD_DEFS)}
    with TEMPLATE_LOCK:
        settings = read_template_settings()
        catalog_settings = settings.setdefault("fieldCatalog", {})
        overrides = catalog_settings.setdefault("overrides", {})
        custom = catalog_settings.setdefault("custom", [])
        if key in builtin_keys:
            current = dict(overrides.get(f"{category}:{key}", {}))
            current["enabled"] = False
            overrides[f"{category}:{key}"] = current
            changed = True
        else:
            original = len(custom)
            custom[:] = [field for field in custom if not (field.get("category") == category and field.get("key") == key)]
            changed = len(custom) != original
        if changed:
            write_template_settings(settings)
    return changed


def create_full_backup() -> tuple[bytes, str, dict]:
    """Create a portable backup containing every user-owned data file."""
    output = io.BytesIO()
    included_files = []
    with zipfile.ZipFile(output, "w", compression=zipfile.ZIP_DEFLATED, compresslevel=6) as archive:
        for name in sorted(BACKUP_ROOT_FILES):
            source = DATA_DIR / name
            if source.is_file():
                archive.write(source, name)
                included_files.append(name)
        for directory_name in sorted(BACKUP_ROOT_DIRS):
            source_directory = DATA_DIR / directory_name
            if not source_directory.is_dir():
                continue
            for source in sorted(path for path in source_directory.rglob("*") if path.is_file()):
                relative = source.relative_to(DATA_DIR).as_posix()
                archive.write(source, relative)
                included_files.append(relative)
        manifest = {
            "format": "freight-invoice-full-backup",
            "formatVersion": 1,
            "appVersion": APP_VERSION,
            "exportedAt": datetime.now().isoformat(timespec="seconds"),
            "files": included_files,
            "counts": {
                "products": len(read_products()),
                "templates": len(read_custom_templates()),
                "history": len(read_export_history()),
                "hasDraft": read_invoice_draft() is not None,
            },
        }
        archive.writestr("backup_manifest.json", json.dumps(manifest, ensure_ascii=False, indent=2))
    filename = f"货代发票完整备份-{datetime.now().strftime('%Y-%m-%d-%H%M%S')}.zip"
    return output.getvalue(), filename, manifest


def backup_member_path(member_name: str) -> PurePosixPath:
    normalized = PurePosixPath(member_name.replace("\\", "/"))
    if normalized.is_absolute() or ".." in normalized.parts or not normalized.parts:
        raise ValueError("备份中包含不安全的文件路径")
    root = normalized.parts[0]
    if member_name == "backup_manifest.json":
        return normalized
    if root not in BACKUP_ROOT_FILES and root not in BACKUP_ROOT_DIRS and root not in BACKUP_IGNORED_FILES:
        raise ValueError(f"备份中包含未知数据：{root}")
    if root in BACKUP_ROOT_FILES | BACKUP_IGNORED_FILES and len(normalized.parts) != 1:
        raise ValueError(f"备份文件路径不正确：{member_name}")
    return normalized


def validate_staged_backup(stage: Path) -> dict:
    manifest_path = stage / "backup_manifest.json"
    if not manifest_path.is_file():
        raise ValueError("这不是货代发票工作台的完整备份")
    try:
        manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, UnicodeDecodeError) as exc:
        raise ValueError("备份清单已损坏") from exc
    if manifest.get("format") != "freight-invoice-full-backup" or int(number(manifest.get("formatVersion"))) != 1:
        raise ValueError("备份版本不受支持")
    json_expectations = {
        "products.json": list,
        "export_history.json": list,
        "custom_templates.json": list,
        "template_settings.json": dict,
        "invoice_draft.json": dict,
    }
    for name, expected_type in json_expectations.items():
        source = stage / name
        if not source.is_file():
            continue
        try:
            value = json.loads(source.read_text(encoding="utf-8"))
        except (json.JSONDecodeError, UnicodeDecodeError) as exc:
            raise ValueError(f"备份中的 {name} 已损坏") from exc
        if not isinstance(value, expected_type):
            raise ValueError(f"备份中的 {name} 格式不正确")
    records_file = stage / "custom_templates.json"
    if records_file.is_file():
        for record in json.loads(records_file.read_text(encoding="utf-8")):
            stored_file = safe_text(record.get("storedFile")) if isinstance(record, dict) else ""
            if not stored_file or not (stage / "custom_templates" / stored_file).is_file():
                raise ValueError("备份中的模板记录缺少对应的 Excel 原文件")
    return manifest


def restore_full_backup(content: bytes) -> dict:
    if not content or len(content) > BACKUP_MAX_BYTES:
        raise ValueError("备份文件为空或超过 1GB")
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    stage = DATA_DIR / f".restore-{uuid.uuid4().hex}"
    stage.mkdir(parents=True, exist_ok=False)
    try:
        try:
            archive = zipfile.ZipFile(io.BytesIO(content))
        except zipfile.BadZipFile as exc:
            raise ValueError("备份 ZIP 已损坏") from exc
        with archive:
            total_size = sum(member.file_size for member in archive.infolist())
            if total_size > BACKUP_MAX_BYTES:
                raise ValueError("备份解压后超过 1GB")
            for member in archive.infolist():
                if member.is_dir():
                    continue
                relative = backup_member_path(member.filename)
                target = stage.joinpath(*relative.parts)
                target.parent.mkdir(parents=True, exist_ok=True)
                with archive.open(member) as source, target.open("wb") as destination:
                    shutil.copyfileobj(source, destination)
        manifest = validate_staged_backup(stage)
        with HISTORY_LOCK, DRAFT_LOCK, TEMPLATE_LOCK:
            for name in sorted(BACKUP_ROOT_FILES):
                target = DATA_DIR / name
                if target.is_file():
                    target.unlink()
            for name in sorted(BACKUP_ROOT_DIRS):
                target = DATA_DIR / name
                if target.is_dir():
                    shutil.rmtree(target)
            for name in sorted(BACKUP_ROOT_FILES):
                source = stage / name
                if source.is_file():
                    shutil.copy2(source, DATA_DIR / name)
            for name in sorted(BACKUP_ROOT_DIRS):
                source = stage / name
                if source.is_dir():
                    shutil.copytree(source, DATA_DIR / name)
        return {
            "products": len(read_products()),
            "templates": len(read_custom_templates()),
            "history": len(read_export_history()),
            "hasDraft": read_invoice_draft() is not None,
            "exportedAt": manifest.get("exportedAt"),
        }
    finally:
        if stage.is_dir():
            shutil.rmtree(stage, ignore_errors=True)


def disable_builtin_template(template_id: str) -> bool:
    if template_id not in TEMPLATES:
        return False
    with TEMPLATE_LOCK:
        settings = read_template_settings()
        disabled = set(settings.get("disabledBuiltins") or [])
        disabled.add(template_id)
        settings["disabledBuiltins"] = sorted(disabled)
        write_template_settings(settings)
    return True


def template_catalog(include_pending=True) -> list[dict]:
    custom = [export_custom_template_summary(record) for record in read_custom_templates()]
    if not include_pending:
        custom = [record for record in custom if record.get("configured")]
    return custom


def export_custom_template_summary(record: dict) -> dict:
    summary = {key: value for key, value in record.items() if key not in {"mapping", "storedFile", "storedOriginalFile"}}
    mapping = record.get("mapping") or {}
    summary["rowMode"] = safe_text(mapping.get("items", {}).get("rowMode"), "sku")
    summary["fixedFieldKeys"] = list((mapping.get("fixed") or {}).keys())
    summary["itemFieldKeys"] = list((mapping.get("items", {}).get("columns") or {}).keys())
    required = template_required_mapping(record)
    summary["requiredFixedKeys"] = required["fixed"]
    summary["requiredItemKeys"] = required["items"]
    return summary


def template_required_mapping(record: dict) -> dict[str, list[str]]:
    mapping = record.get("mapping") or {}
    configured = mapping.get("required")
    if isinstance(configured, dict):
        return {
            "fixed": [safe_text(key) for key in configured.get("fixed", []) if safe_text(key)],
            "items": [safe_text(key) for key in configured.get("items", []) if safe_text(key)],
        }
    cache_key = (safe_text(record.get("id")), safe_text(record.get("updatedAt")))
    if cache_key in TEMPLATE_REQUIRED_CACHE:
        return copy.deepcopy(TEMPLATE_REQUIRED_CACHE[cache_key])
    try:
        path = custom_template_path(record)
        workbook = load_workbook(path, read_only=True, data_only=False, keep_vba=path.suffix.casefold() == ".xlsm")
        ws = workbook[safe_text(mapping.get("sheet"))] if safe_text(mapping.get("sheet")) in workbook.sheetnames else workbook[workbook.sheetnames[0]]
        required_fixed = []
        for field, address in (mapping.get("fixed") or {}).items():
            try:
                target = ws[safe_text(address)]
            except (KeyError, ValueError):
                continue
            # Freight templates normally place the field label immediately to
            # the left of its value cell (possibly with merged value cells).
            labels = []
            for column in range(target.column - 1, max(0, target.column - 5), -1):
                text = safe_text(ws.cell(target.row, column).value)
                if text:
                    labels.append(text)
                    break
            if any(required_marker(label) for label in labels):
                required_fixed.append(field)
        item_mapping = mapping.get("items") or {}
        header_row = max(1, int(number(item_mapping.get("headerRow"), 1)))
        header_start_row = min(header_row, max(1, int(number(item_mapping.get("headerStartRow"), header_row))))
        required_items = []
        for field, mapped_column in (item_mapping.get("columns") or {}).items():
            column = mapping_column(mapped_column)
            if column and required_marker(" ".join(safe_text(ws.cell(row, column).value) for row in range(header_start_row, header_row + 1))):
                required_items.append(field)
        # Older records did not store required metadata. Service/channel is a
        # shipment-level routing choice, so keep it safe by default; users can
        # explicitly uncheck it and save the template if a carrier allows blank.
        if "service" in (mapping.get("fixed") or {}) and "service" not in required_fixed:
            required_fixed.append("service")
        workbook.close()
        inferred = {"fixed": required_fixed, "items": required_items}
        TEMPLATE_REQUIRED_CACHE[cache_key] = inferred
        return copy.deepcopy(inferred)
    except Exception:
        return {"fixed": [], "items": []}


def custom_template_detail(record: dict) -> dict:
    detail = copy.deepcopy(record)
    detail.setdefault("mapping", {})["required"] = template_required_mapping(record)
    return detail


def template_editor_field_catalog(record: dict) -> dict:
    catalog = field_catalog(include_disabled=True)
    mapping = record.get("mapping") or {}
    mapped = {
        "fixed": set((mapping.get("fixed") or {}).keys()),
        "items": set((mapping.get("items", {}).get("columns") or {}).keys()),
    }
    return {
        category: [field for field in fields if field.get("enabled") or field["key"] in mapped[category]]
        for category, fields in catalog.items()
    }


def custom_template_record(template_id: str) -> dict | None:
    return next((record for record in read_custom_templates() if safe_text(record.get("id")) == template_id), None)


def template_display_name(template_id: str) -> str:
    if template_id in TEMPLATES:
        return TEMPLATES[template_id]["name"]
    record = custom_template_record(template_id)
    return safe_text(record.get("name")) if record else template_id


def custom_template_path(record: dict) -> Path:
    filename = safe_text(record.get("storedFile"))
    if not re.fullmatch(r"tpl-[0-9a-f]{10}\.(xlsx|xlsm)", filename, re.IGNORECASE):
        raise ValueError("自定义模板文件记录无效")
    path = (CUSTOM_TEMPLATES_DIR / filename).resolve()
    if CUSTOM_TEMPLATES_DIR.resolve() not in path.parents:
        raise ValueError("自定义模板路径无效")
    return path


def normalized_label(value) -> str:
    return re.sub(r"[\s:：*（）()\[\]_/\\.-]+", "", safe_text(value).casefold())


def required_marker(value) -> bool:
    text = safe_text(value).casefold()
    return "*" in text or "＊" in text or "必填" in text or "required" in text


def looks_like_template_instruction(value) -> bool:
    text = safe_text(value).casefold()
    label = normalized_label(text)
    return len(label) > 4 and any(marker in text for marker in ("说明", "提示", "请填写", "需要填写", "才需要", "必须", "务必", "否则", "必填", "不承担"))


def matched_field(value, allowed_fields: set[str], aliases_by_field: dict | None = None) -> str | None:
    label = normalized_label(value)
    if not label:
        return None
    best = None
    best_length = 0
    for field, aliases in (aliases_by_field or FIELD_ALIASES).items():
        if field not in allowed_fields:
            continue
        for alias in aliases:
            normalized_alias = normalized_label(alias)
            if normalized_alias and normalized_alias in label and len(normalized_alias) > best_length:
                best = field
                best_length = len(normalized_alias)
    return best


def suggested_value_cell(ws, cell) -> str:
    end_col = cell.column
    for merged in ws.merged_cells.ranges:
        if cell.coordinate in merged:
            end_col = merged.max_col
            break
    return ws.cell(cell.row, end_col + 1).coordinate


def merged_anchor_coordinate(ws, coordinate: str) -> str:
    """Return the writable top-left coordinate for normal or merged cells."""
    cell = ws[safe_text(coordinate).upper()]
    for merged in ws.merged_cells.ranges:
        if cell.coordinate in merged:
            return ws.cell(merged.min_row, merged.min_col).coordinate
    return cell.coordinate


def combined_item_header_mapping(ws, first_row: int, last_row: int, scan_cols: int, item_fields: set[str], aliases_by_field: dict | None = None) -> tuple[dict, list[str]]:
    """Match item columns using up to several stacked header rows plus units."""
    columns = {}
    labels = []
    for col in range(1, scan_cols + 1):
        parts = []
        for row in range(first_row, last_row + 1):
            value = safe_text(ws.cell(row, col).value)
            if value and value not in parts:
                parts.append(value)
        combined = " ".join(parts)
        labels.append(normalized_label(combined))
        field = matched_field(combined, item_fields - set(columns), aliases_by_field)
        if field:
            columns[field] = col
    return columns, labels


def analyze_template_workbook(content: bytes, extension: str) -> dict:
    keep_vba = extension.casefold() == ".xlsm"
    workbook = load_workbook(io.BytesIO(content), data_only=False, keep_vba=keep_vba)
    catalog = field_catalog()
    fixed_fields = {field["key"] for field in catalog["fixed"]}
    item_fields = {field["key"] for field in catalog["items"]}
    fixed_aliases = {field["key"]: [field["label"], *field["aliases"]] for field in catalog["fixed"]}
    item_aliases = {field["key"]: [field["label"], *field["aliases"]] for field in catalog["items"]}
    sheet_analyses = []
    for candidate_ws in workbook.worksheets:
        fixed_mapping = {}
        fixed_sources = {}
        header_candidates = []
        nonempty = 0
        scan_rows = min(candidate_ws.max_row, 100)
        scan_cols = min(candidate_ws.max_column, 80)
        for row in range(1, scan_rows + 1):
            for col in range(1, scan_cols + 1):
                cell = candidate_ws.cell(row, col)
                if cell.value not in (None, ""):
                    nonempty += 1
                if not isinstance(cell.value, str):
                    continue
                fixed_field = None
                if len(normalized_label(cell.value)) <= 50 and not looks_like_template_instruction(cell.value):
                    fixed_field = matched_field(cell.value, fixed_fields - set(fixed_mapping), fixed_aliases)
                if fixed_field and fixed_field not in fixed_mapping:
                    fixed_mapping[fixed_field] = suggested_value_cell(candidate_ws, cell)
                    fixed_sources[fixed_field] = (row, col)
            for depth in range(1, min(3, row) + 1):
                first_row = row - depth + 1
                row_mapping, _ = combined_item_header_mapping(candidate_ws, first_row, row, scan_cols, item_fields, item_aliases)
                if row_mapping:
                    header_candidates.append((len(row_mapping), row, row_mapping, depth))
        best_header_count = max((item[0] for item in header_candidates), default=0)
        score = best_header_count * 1000 + len(fixed_mapping) * 100 + min(nonempty, 99)
        sheet_analyses.append((score, candidate_ws, fixed_mapping, header_candidates, fixed_sources))
    _, ws, fixed_mapping, header_candidates, fixed_sources = max(sheet_analyses, key=lambda item: item[0])
    if header_candidates:
        _, header_row, columns, header_depth = max(header_candidates, key=lambda item: (item[0], -item[3], item[1]))
    else:
        header_row, columns, header_depth = 1, {}, 1
    header_start_row = max(1, header_row - header_depth + 1)
    fixed_mapping = {
        field: coordinate for field, coordinate in fixed_mapping.items()
        if not (header_start_row <= fixed_sources.get(field, (0, 0))[0] <= header_row)
    }
    # A visually large merged input area can sit beside several row labels.
    # It is still only one Excel cell, so retain the first semantic match
    # instead of reporting several fields that would overwrite each other.
    unique_fixed_mapping = {}
    occupied_fixed_targets = set()
    for field, coordinate in fixed_mapping.items():
        anchor = merged_anchor_coordinate(ws, coordinate)
        if anchor in occupied_fixed_targets:
            continue
        unique_fixed_mapping[field] = anchor
        occupied_fixed_targets.add(anchor)
    fixed_mapping = unique_fixed_mapping
    required_fixed = [
        field for field in fixed_mapping
        if required_marker(ws.cell(*fixed_sources.get(field, (0, 0))).value)
    ]
    start_row = header_row + 1
    _, header_labels = combined_item_header_mapping(ws, header_start_row, header_row, min(ws.max_column, 80), item_fields, item_aliases)
    row_mode = "carton" if "boxNumber" in columns and any("单箱" in label for label in header_labels) else "sku"
    required_items = []
    for field, column in columns.items():
        header_text = " ".join(safe_text(ws.cell(row, column).value) for row in range(header_start_row, header_row + 1))
        if required_marker(header_text):
            required_items.append(field)
    reserved_rows = 1
    for candidate_row in range(start_row + 1, min(ws.max_row, start_row + 20) + 1):
        row_labels = [normalized_label(ws.cell(candidate_row, col).value) for col in range(1, min(ws.max_column, 80) + 1)]
        if any(label in {"total", "total:", "合计", "总计"} or label.startswith("total") for label in row_labels if label):
            reserved_rows = max(1, candidate_row - start_row)
            break
    mapping = {
        "sheet": ws.title,
        "fixed": fixed_mapping,
        "required": {"fixed": required_fixed, "items": required_items},
        "items": {
            "headerStartRow": header_start_row,
            "headerRow": header_row,
            "startRow": start_row,
            "reservedRows": reserved_rows,
            "rowMode": row_mode,
            "columns": columns,
            "imageMaxWidth": 72,
            "imageMaxHeight": 50,
        },
    }
    fixed_examples = {}
    for field, coordinate in fixed_mapping.items():
        try:
            value = ws[coordinate].value
            if value not in (None, ""):
                fixed_examples[field] = safe_text(value)
        except Exception:
            continue
    result = {
        "sheetNames": workbook.sheetnames,
        "selectedSheet": ws.title,
        "maxRow": ws.max_row,
        "maxCol": ws.max_column,
        "mapping": mapping,
        "fixedExamples": fixed_examples,
        "detectedFixed": len(fixed_mapping),
        "detectedItemColumns": len(columns),
    }
    workbook.close()
    return result


def convert_xls_with_excel(content: bytes, filename: str) -> bytes:
    """Convert legacy Excel through desktop Excel when available, preserving fidelity."""
    import pythoncom
    import win32com.client

    pythoncom.CoInitialize()
    excel = None
    workbook = None
    try:
        DATA_DIR.mkdir(parents=True, exist_ok=True)
        with tempfile.TemporaryDirectory(prefix="template-convert-", dir=DATA_DIR) as folder:
            input_path = Path(folder) / "source.xls"
            output_path = Path(folder) / "converted.xlsx"
            input_path.write_bytes(content)
            excel = win32com.client.DispatchEx("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False
            excel.AskToUpdateLinks = False
            workbook = excel.Workbooks.Open(str(input_path), UpdateLinks=0, ReadOnly=True)
            workbook.SaveAs(str(output_path), FileFormat=51)
            workbook.Close(False)
            workbook = None
            return output_path.read_bytes()
    finally:
        if workbook is not None:
            try:
                workbook.Close(False)
            except Exception:
                pass
        if excel is not None:
            try:
                excel.Quit()
            except Exception:
                pass
        pythoncom.CoUninitialize()


def convert_xls_with_powershell(content: bytes) -> bytes:
    """Use installed desktop Excel through Windows PowerShell without pywin32."""
    script = BUNDLE_ROOT / "tools" / "convert_xls_to_xlsx.ps1"
    if not script.is_file():
        raise FileNotFoundError("Excel 转换脚本不存在")
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    with tempfile.TemporaryDirectory(prefix="template-excel-", dir=DATA_DIR) as folder:
        input_path = Path(folder) / "source.xls"
        output_path = Path(folder) / "converted.xlsx"
        input_path.write_bytes(content)
        completed = subprocess.run(
            [
                "powershell.exe", "-NoProfile", "-ExecutionPolicy", "Bypass", "-File", str(script),
                "-InputPath", str(input_path), "-OutputPath", str(output_path),
            ],
            cwd=APP_ROOT,
            capture_output=True,
            text=True,
            timeout=90,
            creationflags=getattr(subprocess, "CREATE_NO_WINDOW", 0),
        )
        if completed.returncode != 0 or not output_path.is_file():
            raise RuntimeError(completed.stderr.strip() or completed.stdout.strip() or "Excel 转换失败")
        return output_path.read_bytes()


def xls_color(book, index) -> str | None:
    if index in (None, 64, 65):
        return None
    rgb = book.colour_map.get(index)
    return "FF%02X%02X%02X" % rgb if rgb else None


def convert_xls_with_xlrd(content: bytes) -> bytes:
    """Portable fallback for computers without Excel; formulas become cached values."""
    import xlrd

    source = xlrd.open_workbook(file_contents=content, formatting_info=True)
    target = Workbook()
    target.remove(target.active)
    border_styles = {
        0: None, 1: "thin", 2: "medium", 3: "dashed", 4: "dotted", 5: "thick",
        6: "double", 7: "hair", 8: "mediumDashed", 9: "dashDot", 10: "mediumDashDot",
        11: "dashDotDot", 12: "mediumDashDotDot", 13: "slantDashDot",
    }
    horizontal = {1: "left", 2: "center", 3: "right", 4: "fill", 5: "justify", 6: "centerContinuous", 7: "distributed"}
    vertical = {0: "top", 1: "center", 2: "bottom", 3: "justify", 4: "distributed"}
    for source_sheet in source.sheets():
        title = source_sheet.name[:31] or "Sheet"
        sheet = target.create_sheet(title)
        for col_index, info in source_sheet.colinfo_map.items():
            sheet.column_dimensions[get_column_letter(col_index + 1)].width = max(0.1, info.width / 256)
            sheet.column_dimensions[get_column_letter(col_index + 1)].hidden = bool(info.hidden)
        for row_index, info in source_sheet.rowinfo_map.items():
            sheet.row_dimensions[row_index + 1].height = info.height / 20
            sheet.row_dimensions[row_index + 1].hidden = bool(info.hidden)
        for row_index in range(source_sheet.nrows):
            for col_index in range(source_sheet.ncols):
                old_cell = source_sheet.cell(row_index, col_index)
                value = old_cell.value
                if old_cell.ctype == xlrd.XL_CELL_DATE:
                    value = xlrd.xldate.xldate_as_datetime(value, source.datemode)
                elif old_cell.ctype == xlrd.XL_CELL_BOOLEAN:
                    value = bool(value)
                elif old_cell.ctype in (xlrd.XL_CELL_EMPTY, xlrd.XL_CELL_BLANK):
                    value = None
                elif old_cell.ctype == xlrd.XL_CELL_ERROR:
                    value = xlrd.error_text_from_code.get(value, "#ERROR!")
                cell = sheet.cell(row_index + 1, col_index + 1, value)
                if old_cell.xf_index >= len(source.xf_list):
                    continue
                xf = source.xf_list[old_cell.xf_index]
                old_font = source.font_list[xf.font_index]
                cell.font = Font(
                    name=old_font.name or "Calibri", size=(old_font.height or 220) / 20,
                    bold=bool(old_font.bold), italic=bool(old_font.italic),
                    underline="single" if old_font.underline_type else None,
                    color=xls_color(source, old_font.colour_index),
                )
                fill_color = xls_color(source, xf.background.pattern_colour_index)
                if fill_color and xf.background.fill_pattern:
                    cell.fill = PatternFill("solid", fgColor=fill_color)
                cell.alignment = Alignment(
                    horizontal=horizontal.get(xf.alignment.hor_align),
                    vertical=vertical.get(xf.alignment.vert_align),
                    wrap_text=bool(xf.alignment.text_wrapped),
                    text_rotation=xf.alignment.rotation if 0 <= xf.alignment.rotation <= 180 else 0,
                )
                def side(style, color_index):
                    return Side(style=border_styles.get(style), color=xls_color(source, color_index))
                old_border = xf.border
                cell.border = Border(
                    left=side(old_border.left_line_style, old_border.left_colour_index),
                    right=side(old_border.right_line_style, old_border.right_colour_index),
                    top=side(old_border.top_line_style, old_border.top_colour_index),
                    bottom=side(old_border.bottom_line_style, old_border.bottom_colour_index),
                )
                format_record = source.format_map.get(xf.format_key)
                if format_record and format_record.format_str:
                    cell.number_format = format_record.format_str
        for row_low, row_high, col_low, col_high in source_sheet.merged_cells:
            if row_high > row_low and col_high > col_low:
                sheet.merge_cells(start_row=row_low + 1, end_row=row_high, start_column=col_low + 1, end_column=col_high)
    output = io.BytesIO()
    target.save(output)
    target.close()
    source.release_resources()
    return output.getvalue()


def normalize_template_upload(filename: str, content: bytes) -> tuple[str, bytes, str]:
    extension = Path(filename).suffix.casefold()
    if extension != ".xls":
        return extension, content, "original"
    if not content.startswith(b"\xD0\xCF\x11\xE0\xA1\xB1\x1A\xE1"):
        raise ValueError("文件不是有效的旧版 Excel .xls 工作簿")
    try:
        return ".xlsx", convert_xls_with_excel(content, filename), "excel"
    except Exception:
        try:
            return ".xlsx", convert_xls_with_powershell(content), "excel"
        except Exception:
            try:
                return ".xlsx", convert_xls_with_xlrd(content), "compatible"
            except Exception as exc:
                raise ValueError("无法读取该 .xls 文件，请尝试用 Excel 另存为 .xlsx") from exc


def validate_template_upload(filename: str, content: bytes) -> tuple[str, bytes, dict, str]:
    extension = Path(filename).suffix.casefold()
    if extension not in {".xls", ".xlsx", ".xlsm"}:
        raise ValueError("第二版支持 .xls、.xlsx 和 .xlsm 模板")
    if not content or len(content) > 20 * 1024 * 1024:
        raise ValueError("模板文件为空或超过 20MB")
    normalized_extension, normalized_content, conversion_mode = normalize_template_upload(filename, content)
    try:
        with zipfile.ZipFile(io.BytesIO(normalized_content)) as archive:
            if sum(info.file_size for info in archive.infolist()) > 120 * 1024 * 1024:
                raise ValueError("模板解压后内容过大")
    except zipfile.BadZipFile as exc:
        raise ValueError("文件不是有效的 Excel 工作簿") from exc
    return normalized_extension, normalized_content, analyze_template_workbook(normalized_content, normalized_extension), conversion_mode


def create_custom_template(name: str, filename: str, content: bytes) -> dict:
    extension, normalized_content, analysis, conversion_mode = validate_template_upload(filename, content)
    template_id = f"tpl-{uuid.uuid4().hex[:10]}"
    stored_file = f"{template_id}{extension}"
    stored_original_file = f"{template_id}.xls" if Path(filename).suffix.casefold() == ".xls" else ""
    record = {
        "id": template_id,
        "name": safe_text(name) or Path(filename).stem,
        "description": f"自定义模板 · {Path(filename).name}",
        "originalFilename": Path(filename).name,
        "sourceFormat": Path(filename).suffix.casefold(),
        "conversionMode": conversion_mode,
        "storedFile": stored_file,
        "builtIn": False,
        "configured": False,
        "createdAt": datetime.now().isoformat(timespec="seconds"),
        "sheetNames": analysis["sheetNames"],
        "detectedFixed": analysis["detectedFixed"],
        "detectedItemColumns": analysis["detectedItemColumns"],
        "mapping": analysis["mapping"],
        "fixedExamples": analysis["fixedExamples"],
        "analyzerVersion": 3,
    }
    if stored_original_file:
        record["storedOriginalFile"] = stored_original_file
    with TEMPLATE_LOCK:
        CUSTOM_TEMPLATES_DIR.mkdir(parents=True, exist_ok=True)
        destination = CUSTOM_TEMPLATES_DIR / stored_file
        temporary = destination.with_suffix(destination.suffix + ".tmp")
        temporary.write_bytes(normalized_content)
        temporary.replace(destination)
        if stored_original_file:
            original_destination = CUSTOM_TEMPLATES_DIR / stored_original_file
            original_temporary = original_destination.with_suffix(".xls.tmp")
            original_temporary.write_bytes(content)
            original_temporary.replace(original_destination)
        records = read_custom_templates()
        records.append(record)
        write_custom_templates(records)
    return record


def convert_builtin_to_custom(template_id: str) -> dict:
    if template_id not in TEMPLATES:
        raise ValueError("系统模板不存在")
    source = TEMPLATES[template_id]["file"]
    record = create_custom_template(TEMPLATES[template_id]["name"], source.name, source.read_bytes())
    with TEMPLATE_LOCK:
        records = read_custom_templates()
        stored = next((item for item in records if item.get("id") == record["id"]), None)
        if stored is not None:
            stored["description"] = f"通用可编辑模板 · 原 {TEMPLATES[template_id]['name']}"
            stored["migratedFromBuiltin"] = template_id
            write_custom_templates(records)
            record = stored
    disable_builtin_template(template_id)
    return record


def update_custom_template_mapping(template_id: str, payload: dict) -> dict:
    with TEMPLATE_LOCK:
        records = read_custom_templates()
        record = next((item for item in records if safe_text(item.get("id")) == template_id), None)
        if record is None:
            raise ValueError("自定义模板不存在")
        mapping = payload.get("mapping")
        if not isinstance(mapping, dict) or not safe_text(mapping.get("sheet")):
            raise ValueError("模板映射格式不正确")
        item_mapping = mapping.get("items", {})
        if int(number(item_mapping.get("startRow"))) < 1:
            raise ValueError("请设置商品明细起始行")
        record["name"] = safe_text(payload.get("name")) or record["name"]
        record["description"] = safe_text(payload.get("description")) or record["description"]
        record["mapping"] = mapping
        record["configured"] = True
        record["updatedAt"] = datetime.now().isoformat(timespec="seconds")
        write_custom_templates(records)
    return record


def reanalyze_custom_template(template_id: str) -> dict:
    with TEMPLATE_LOCK:
        records = read_custom_templates()
        record = next((item for item in records if safe_text(item.get("id")) == template_id), None)
        if record is None:
            raise ValueError("自定义模板不存在")
        template_path = custom_template_path(record)
        analysis = analyze_template_workbook(template_path.read_bytes(), template_path.suffix.casefold())
        record["sheetNames"] = analysis["sheetNames"]
        record["detectedFixed"] = analysis["detectedFixed"]
        record["detectedItemColumns"] = analysis["detectedItemColumns"]
        record["mapping"] = analysis["mapping"]
        record["fixedExamples"] = analysis["fixedExamples"]
        record["analyzerVersion"] = 3
        record["configured"] = False
        record["updatedAt"] = datetime.now().isoformat(timespec="seconds")
        write_custom_templates(records)
    return record


def unlink_template_file(path: Path) -> None:
    try:
        path.unlink(missing_ok=True)
    except OSError:
        # Windows can briefly retain an Excel ZIP handle immediately after an
        # export. The catalog deletion should still succeed; retry after the
        # request has released all workbook objects.
        def retry():
            try:
                path.unlink(missing_ok=True)
            except OSError:
                pass
        timer = threading.Timer(1.0, retry)
        timer.daemon = True
        timer.start()


def delete_custom_template(template_id: str) -> bool:
    with TEMPLATE_LOCK:
        records = read_custom_templates()
        record = next((item for item in records if safe_text(item.get("id")) == template_id), None)
        if record is None:
            return False
        kept = [item for item in records if safe_text(item.get("id")) != template_id]
        write_custom_templates(kept)
        unlink_template_file(custom_template_path(record))
        original_filename = safe_text(record.get("storedOriginalFile"))
        if re.fullmatch(r"tpl-[0-9a-f]{10}\.xls", original_filename, re.IGNORECASE):
            unlink_template_file(CUSTOM_TEMPLATES_DIR / original_filename)
    return True


def delete_custom_templates(template_ids: list) -> int:
    selected = {safe_text(value) for value in template_ids if re.fullmatch(r"tpl-[0-9a-f]{10}", safe_text(value), re.IGNORECASE)}
    if not selected:
        return 0
    with TEMPLATE_LOCK:
        records = read_custom_templates()
        removed = [record for record in records if safe_text(record.get("id")) in selected]
        if not removed:
            return 0
        write_custom_templates([record for record in records if safe_text(record.get("id")) not in selected])
        for record in removed:
            for filename in (safe_text(record.get("storedFile")), safe_text(record.get("storedOriginalFile"))):
                if not re.fullmatch(r"tpl-[0-9a-f]{10}\.(?:xls|xlsx|xlsm)", filename, re.IGNORECASE):
                    continue
                unlink_template_file(CUSTOM_TEMPLATES_DIR / filename)
    return len(removed)


def read_export_history() -> list:
    try:
        data = json.loads(HISTORY_FILE.read_text(encoding="utf-8"))
        return data if isinstance(data, list) else []
    except (FileNotFoundError, json.JSONDecodeError, OSError):
        return []


def write_export_history(records: list) -> None:
    DATA_DIR.mkdir(exist_ok=True)
    temporary = HISTORY_FILE.with_suffix(".json.tmp")
    temporary.write_text(json.dumps(records, ensure_ascii=False, indent=2), encoding="utf-8")
    temporary.replace(HISTORY_FILE)


def export_history_summary(record: dict) -> dict:
    return {key: value for key, value in record.items() if key != "payload"}


def read_invoice_draft() -> dict | None:
    try:
        data = json.loads(DRAFT_FILE.read_text(encoding="utf-8"))
        return data if isinstance(data, dict) else None
    except (FileNotFoundError, json.JSONDecodeError, OSError):
        return None


def write_invoice_draft(payload: dict) -> dict:
    record = {"updatedAt": datetime.now().isoformat(timespec="seconds"), "payload": copy.deepcopy(payload)}
    with DRAFT_LOCK:
        DATA_DIR.mkdir(exist_ok=True)
        temporary = DRAFT_FILE.with_suffix(".json.tmp")
        temporary.write_text(json.dumps(record, ensure_ascii=False, indent=2), encoding="utf-8")
        temporary.replace(DRAFT_FILE)
    return {"updatedAt": record["updatedAt"]}


def clear_invoice_draft() -> None:
    with DRAFT_LOCK:
        try:
            DRAFT_FILE.unlink()
        except FileNotFoundError:
            pass


def history_file_path(record_id: str) -> Path:
    if not re.fullmatch(r"[0-9]{14}-[0-9a-f]{8}", record_id):
        raise ValueError("历史记录编号无效")
    return HISTORY_FILES_DIR / f"{record_id}.xlsx"


def save_export_history(payload: dict, filename: str, content: bytes) -> dict:
    shipment = payload.get("shipment", {})
    items = payload.get("items", [])
    cartons = sum(max(1, int(number(item.get("cartons"), 1))) for item in items)
    totals = {}
    for item in items:
        currency = safe_text(item.get("currency"), "USD").upper()
        item_cartons = max(1, int(number(item.get("cartons"), 1)))
        totals[currency] = totals.get(currency, 0) + number(item.get("quantity")) * number(item.get("unitPrice")) * item_cartons
    template_id = safe_text(payload.get("templateId"))
    template_name = template_display_name(template_id)
    record_id = f"{datetime.now().strftime('%Y%m%d%H%M%S')}-{uuid.uuid4().hex[:8]}"
    record = {
        "id": record_id,
        "createdAt": datetime.now().isoformat(timespec="seconds"),
        "templateId": template_id,
        "templateName": template_name,
        "filename": filename,
        "fbaNumber": safe_text(shipment.get("fbaNumber")),
        "amazonReferenceId": safe_text(shipment.get("amazonReferenceId")),
        "receiver": safe_text(shipment.get("receiverCompany") or shipment.get("receiverName")),
        "itemCount": len(items),
        "cartonCount": cartons,
        "totals": totals,
        "fileStored": True,
        "fileSize": len(content),
        "payload": copy.deepcopy(payload),
    }
    with HISTORY_LOCK:
        HISTORY_FILES_DIR.mkdir(parents=True, exist_ok=True)
        stored_file = history_file_path(record_id)
        temporary_file = stored_file.with_suffix(".xlsx.tmp")
        temporary_file.write_bytes(content)
        temporary_file.replace(stored_file)
        records = read_export_history()
        records.insert(0, record)
        removed = records[100:]
        write_export_history(records[:100])
        for old_record in removed:
            try:
                history_file_path(safe_text(old_record.get("id"))).unlink()
            except (FileNotFoundError, ValueError, OSError):
                pass
    return export_history_summary(record)


def delete_export_history(record_id: str) -> bool:
    with HISTORY_LOCK:
        records = read_export_history()
        kept = [record for record in records if safe_text(record.get("id")) != record_id]
        if len(kept) == len(records):
            return False
        write_export_history(kept)
        try:
            history_file_path(record_id).unlink()
        except FileNotFoundError:
            pass
    return True


def delete_export_histories(record_ids: list) -> int:
    selected = {safe_text(value) for value in record_ids if re.fullmatch(r"\d{14}-[0-9a-f]{8}", safe_text(value), re.IGNORECASE)}
    if not selected:
        return 0
    with HISTORY_LOCK:
        records = read_export_history()
        removed = [record for record in records if safe_text(record.get("id")) in selected]
        if not removed:
            return 0
        write_export_history([record for record in records if safe_text(record.get("id")) not in selected])
        for record in removed:
            try:
                history_file_path(safe_text(record.get("id"))).unlink()
            except FileNotFoundError:
                pass
    return len(removed)


def number(value, default=0.0) -> float:
    try:
        if value in (None, ""):
            return default
        return float(value)
    except (TypeError, ValueError):
        return default


def optional_number(value):
    if value in (None, ""):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def product_weight_text(item: dict) -> str:
    value = optional_number(item.get("productWeight"))
    unit = safe_text(item.get("weightUnit"), "kg").lower()
    if value is None:
        value = optional_number(item.get("unitNetWeight"))
        unit = "kg"
    if value is None:
        return ""
    unit = unit if unit in {"g", "kg", "oz", "lb"} else "kg"
    return f"{value:g} {unit}"


def clean_filename(value: str) -> str:
    value = re.sub(r'[<>:"/\\|?*\x00-\x1f]', "-", safe_text(value))
    return value.strip(" .-")[:80] or "invoice"


def copy_row_style(ws, source_row: int, target_row: int, min_col: int, max_col: int) -> None:
    ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height
    ws.row_dimensions[target_row].hidden = ws.row_dimensions[source_row].hidden
    for col in range(min_col, max_col + 1):
        source = ws.cell(source_row, col)
        target = ws.cell(target_row, col)
        if source.has_style:
            target._style = copy.copy(source._style)
        if source.number_format:
            target.number_format = source.number_format
        if source.alignment:
            target.alignment = copy.copy(source.alignment)
        if source.protection:
            target.protection = copy.copy(source.protection)


def image_bytes(data_url: str) -> io.BytesIO | None:
    if not data_url or not data_url.startswith("data:image/"):
        return None
    try:
        _, payload = data_url.split(",", 1)
        return io.BytesIO(base64.b64decode(payload))
    except (ValueError, TypeError, base64.binascii.Error):
        return None


def add_cell_image(ws, cell: str, data_url: str, max_width=90, max_height=55) -> None:
    stream = image_bytes(data_url)
    if stream is None:
        return
    try:
        image = ExcelImage(stream)
        ratio = min(max_width / max(image.width, 1), max_height / max(image.height, 1), 1)
        image.width = int(image.width * ratio)
        image.height = int(image.height * ratio)
        target = ws[cell]
        column_letter = get_column_letter(target.column)
        column_width = ws.column_dimensions[column_letter].width or 8.43
        # Excel character-width units are approximately seven pixels plus cell padding.
        cell_width_pixels = max(1, int(column_width * 7 + 5))
        row_height_points = ws.row_dimensions[target.row].height or ws.sheet_format.defaultRowHeight or 15
        cell_height_pixels = max(1, int(points_to_pixels(row_height_points)))
        horizontal_offset = max(0, (cell_width_pixels - image.width) // 2)
        vertical_offset = max(0, (cell_height_pixels - image.height) // 2)
        image.anchor = OneCellAnchor(
            _from=AnchorMarker(
                col=target.column - 1,
                row=target.row - 1,
                colOff=pixels_to_EMU(horizontal_offset),
                rowOff=pixels_to_EMU(vertical_offset),
            ),
            ext=XDRPositiveSize2D(
                cx=pixels_to_EMU(image.width),
                cy=pixels_to_EMU(image.height),
            ),
        )
        ws.add_image(image)
    except Exception:
        return


def joined_bilingual(chinese: str, english: str) -> str:
    values = [safe_text(chinese), safe_text(english)]
    return " / ".join(item for item in values if item)


def parsed_shipping_address(value: str) -> dict:
    text = safe_text(value).replace("\r", "")
    result = {}
    location = re.search(
        r"(?:^|\n)\s*([^\n,]+?)\s*,\s*([A-Z]{2})\s+(\d{5}(?:-\d{4})?)(?:\s|$)",
        text,
        re.IGNORECASE,
    )
    if location:
        result.update({
            "city": location.group(1).strip().upper(),
            "state": location.group(2).upper(),
            "postalCode": location.group(3),
            "countryCode": "US",
        })
    known_codes = {"US", "GB", "DE", "FR", "IT", "ES", "PL", "CA", "MX", "AU", "JP", "NL", "BE", "SE", "NO", "DK", "AT", "CZ", "CH", "IE", "PT"}
    for line in reversed([part.strip() for part in text.split("\n") if part.strip()]):
        country = re.match(r"^([A-Z]{2})(?:\s*\(|\s*$)", line, re.IGNORECASE)
        if country and country.group(1).upper() in known_codes:
            result["countryCode"] = country.group(1).upper()
            break
    return result


def expand_yuntuo_cartons(items: list, shipment: dict) -> list:
    requested_box_count = max(1, int(number(shipment.get("boxCount"), 1)))
    fba_number = re.sub(r"U\d{6}$", "", safe_text(shipment.get("fbaNumber")), flags=re.IGNORECASE)
    reference_id = safe_text(shipment.get("amazonReferenceId"))
    expanded = []
    sequence = 1
    for item_index, source in enumerate(items):
        item = dict(source)
        item_box_count = max(1, int(number(item.get("cartons"), 1)))
        if len(items) == 1:
            item_box_count = requested_box_count
        # YunTuo rows describe one carton. Product quantity and weights are
        # therefore per-carton values and must be repeated, not averaged.
        carton_quantity = optional_number(item.get("quantity"))
        carton_gross_weight = optional_number(item.get("grossWeight"))
        carton_net_weight = optional_number(item.get("netWeight"))
        unit_weight_text = product_weight_text(item)
        for carton_index in range(item_box_count):
            row = dict(item)
            row.update({
                "boxNumber": f"{fba_number}U{sequence:06d}",
                "quantity": carton_quantity,
                "grossWeight": carton_gross_weight,
                "netWeight": carton_net_weight,
                "unitWeightText": unit_weight_text,
                "poNumber": reference_id,
                "appointmentWindow": safe_text(item.get("appointmentWindow")) or safe_text(shipment.get("appointmentWindow")),
                "cartons": 1,
            })
            expanded.append(row)
            sequence += 1
    return expanded


def export_ruichi(payload: dict) -> tuple[bytes, str]:
    template = TEMPLATES["ruichi"]["file"]
    workbook = load_workbook(template)
    ws = workbook[workbook.sheetnames[0]]
    shipment = payload.get("shipment", {})
    items = payload.get("items", []) or [{}]

    ws["E3"] = safe_text(shipment.get("receiverCompany"), "美国亚马逊")
    ws["E4"] = safe_text(shipment.get("receiverName"), "美国亚马逊")
    address_parts = [
        safe_text(shipment.get("address1")),
        safe_text(shipment.get("address2")),
        " ".join(filter(None, [safe_text(shipment.get("city")), safe_text(shipment.get("state")), safe_text(shipment.get("postalCode"))])),
        safe_text(shipment.get("countryCode")),
    ]
    address = "\n".join(part for part in address_parts if part)
    contact = "  ".join(filter(None, [safe_text(shipment.get("phone")), safe_text(shipment.get("email"))]))
    ws["E5"] = "\n".join(filter(None, [address, contact]))
    ws["P3"] = safe_text(shipment.get("fbaNumber"))
    ws["P5"] = safe_text(shipment.get("amazonReferenceId"))

    start_row = 10
    original_total_row = 13
    if len(items) > 3:
        extra = len(items) - 3
        ws.insert_rows(original_total_row, amount=extra)
        for row in range(original_total_row, original_total_row + extra):
            copy_row_style(ws, 12, row, 1, 18)
    total_row = start_row + max(len(items), 3)

    for row in range(start_row, total_row):
        for col in range(1, 19):
            ws.cell(row, col).value = None

    currencies = {safe_text(item.get("currency"), "USD").upper() for item in items}
    currency_label = next(iter(currencies)) if len(currencies) == 1 else "MIXED"
    ws["E9"] = currency_label
    ws["F9"] = currency_label

    for index, item in enumerate(items):
        row = start_row + index
        quantity = number(item.get("quantity"))
        unit_price = number(item.get("unitPrice"))
        carton_length = number(item.get("cartonLength"))
        carton_width = number(item.get("cartonWidth"))
        carton_height = number(item.get("cartonHeight"))
        cartons = max(1, number(item.get("cartons"), 1))
        total_quantity = quantity * cartons
        cbm = number(item.get("cbm")) or carton_length * carton_width * carton_height * cartons / 1_000_000
        values = {
            # 锐驰的“单号”是整票 FBA 号；每个 SKU 行使用同一个值，不追加箱序号。
            1: safe_text(shipment.get("fbaNumber")),
            2: safe_text(item.get("nameZh")),
            3: safe_text(item.get("nameEn")),
            4: total_quantity,
            5: unit_price,
            6: total_quantity * unit_price,
            7: cartons,
            8: number(item.get("grossWeight")),
            9: number(item.get("netWeight")),
            11: cbm,
            12: safe_text(item.get("hsCode")),
            13: safe_text(item.get("brand"), "无"),
            14: safe_text(item.get("brandType"), "无品牌"),
            15: safe_text(item.get("model"), "无"),
            16: joined_bilingual(item.get("materialZh"), item.get("materialEn")),
            17: joined_bilingual(item.get("purposeZh"), item.get("purposeEn")),
        }
        for col, value in values.items():
            ws.cell(row, col).value = value
        ws.row_dimensions[row].height = max(ws.row_dimensions[row].height or 15, 48)
        add_cell_image(ws, f"R{row}", safe_text(item.get("image")))

    ws[f"B{total_row}"] = "TOTAL:"
    for col in (4, 6, 7, 8, 9, 11):
        letter = ws.cell(1, col).column_letter
        ws.cell(total_row, col).value = f"=SUM({letter}{start_row}:{letter}{start_row + len(items) - 1})"

    output = io.BytesIO()
    workbook.save(output)
    name = clean_filename(shipment.get("fbaNumber") or shipment.get("orderNumber") or "锐驰发票")
    return output.getvalue(), f"{name}-锐驰.xlsx"


def export_yuntuo(payload: dict) -> tuple[bytes, str]:
    template = TEMPLATES["yuntuo"]["file"]
    workbook = load_workbook(template)
    ws = workbook["发票模版"]
    # The converted legacy .xls contains two broken 20x15 image objects at X26.
    # Remove template artifacts before adding real product images.
    ws._images = []
    shipment = dict(payload.get("shipment", {}))
    source_items = payload.get("items", []) or [{}]
    parsed_address = parsed_shipping_address("\n".join(filter(None, [
        safe_text(shipment.get("address1")), safe_text(shipment.get("address2")),
    ])))
    for key, value in parsed_address.items():
        if not safe_text(shipment.get(key)):
            shipment[key] = value
    items = expand_yuntuo_cartons(source_items, shipment)
    shipment["boxCount"] = len(items)

    base_mapping = {
        "B1": "service",
        "B2": "receiverName",
        "B3": "receiverCompany",
        "B4": "address1",
        "B5": "address2",
        "B6": "city",
        "B7": "state",
        "B8": "postalCode",
        "B9": "countryCode",
        "B10": "phone",
        "B11": "email",
        "B12": "fbaNumber",
        "B13": "amazonReferenceId",
        "B14": "boxCount",
        "B15": "hasBattery",
        "B16": "hasMagnet",
        "B17": "currency",
        "B18": "customsMode",
        "B19": "vatNumber",
        "B20": "eori",
        "B21": "vatName",
        "B22": "vatAddress",
        "B23": "notes",
    }
    defaults = {
        "receiverName": "Amazon Fulfillment Center",
        "hasBattery": "否",
        "hasMagnet": "否",
        "currency": "USD",
        "customsMode": "买单报关",
    }
    for cell, key in base_mapping.items():
        value = shipment.get(key, defaults.get(key, ""))
        # The source template contains example shipment data. Always overwrite
        # editable cells, including with blanks, so examples never leak into exports.
        ws[cell] = value if value not in (None, "") else None

    start_row = 25
    if len(items) > 1:
        ws.insert_rows(start_row + 1, amount=len(items) - 1)
        for row in range(start_row + 1, start_row + len(items)):
            copy_row_style(ws, start_row, row, 1, 25)

    for index, item in enumerate(items):
        row = start_row + index
        quantity = optional_number(item.get("quantity"))
        unit_price = optional_number(item.get("unitPrice"))
        sale_price = optional_number(item.get("salePrice"))
        if sale_price in (None, 0):
            sale_price = unit_price
        unit_weight_text = safe_text(item.get("unitWeightText")) or product_weight_text(item)
        values = {
            1: safe_text(item.get("boxNumber")),
            2: optional_number(item.get("grossWeight")),
            3: optional_number(item.get("netWeight")),
            4: unit_weight_text,
            # Only the carrier-facing alias may leave the app. Never fall back to the internal SKU.
            5: safe_text(item.get("exportSku")),
            6: safe_text(item.get("nameEn")),
            7: safe_text(item.get("nameZh")),
            8: unit_price,
            9: quantity,
            10: f"=I{row}*H{row}",
            11: optional_number(item.get("cartonLength")),
            12: optional_number(item.get("cartonWidth")),
            13: optional_number(item.get("cartonHeight")),
            14: joined_bilingual(item.get("materialZh"), item.get("materialEn")),
            15: joined_bilingual(item.get("purposeZh"), item.get("purposeEn")),
            16: safe_text(item.get("hsCode")),
            17: safe_text(item.get("poNumber")),
            18: safe_text(item.get("brand")),
            19: safe_text(item.get("model")),
            20: "",
            21: safe_text(item.get("hasMagnet")),
            22: safe_text(item.get("hasBattery")),
            23: safe_text(item.get("batteryInfo")),
            24: sale_price,
            25: safe_text(item.get("saleUrl")),
        }
        for col, value in values.items():
            ws.cell(row, col).value = value
        ws.row_dimensions[row].height = max(ws.row_dimensions[row].height or 15, 55)
        add_cell_image(ws, f"T{row}", safe_text(item.get("image")), max_width=72, max_height=50)

    output = io.BytesIO()
    workbook.save(output)
    name = clean_filename(shipment.get("fbaNumber") or shipment.get("orderNumber") or "云拓发票")
    return output.getvalue(), f"{name}-云拓.xlsx"


def normalized_address_part(value) -> str:
    return re.sub(r"[^0-9a-z\u4e00-\u9fff]+", "", safe_text(value).casefold())


def custom_full_address(shipment: dict) -> tuple[str, str]:
    """Build lossless multi-line and compact addresses from split form fields."""
    city = safe_text(shipment.get("city"))
    state = safe_text(shipment.get("state"))
    postal_code = safe_text(shipment.get("postalCode"))
    location = " ".join(part for part in [state, postal_code] if part)
    if city and location:
        location = f"{city}, {location}"
    elif city:
        location = city

    candidates = [
        safe_text(shipment.get("warehouseCode")),
        safe_text(shipment.get("address1")),
        safe_text(shipment.get("address2")),
        safe_text(shipment.get("address3")),
        location,
        safe_text(shipment.get("countryCode")),
    ]
    lines: list[str] = []
    normalized_lines: list[str] = []
    for candidate in candidates:
        for raw_line in re.split(r"[\r\n]+", candidate):
            line = raw_line.strip(" ,")
            normalized = normalized_address_part(line)
            if not normalized:
                continue
            # Preserve the user's original address while avoiding repeated
            # warehouse/locality lines when a pasted address was also parsed.
            combined = "".join(normalized_lines)
            if normalized in normalized_lines or (len(normalized) >= 4 and normalized in combined):
                continue
            lines.append(line)
            normalized_lines.append(normalized)
    return "\n".join(lines), ", ".join(lines)


def custom_fixed_values(payload: dict) -> dict:
    shipment = payload.get("shipment", {})
    items = payload.get("items", [])
    currencies = {safe_text(item.get("currency"), "USD").upper() for item in items}
    total_cartons = sum(max(1, int(number(item.get("cartons"), 1))) for item in items)
    total_quantity = sum(number(item.get("quantity")) * max(1, int(number(item.get("cartons"), 1))) for item in items)
    total_gross = sum(number(item.get("grossWeight")) * max(1, int(number(item.get("cartons"), 1))) for item in items)
    total_net = sum(number(item.get("netWeight")) * max(1, int(number(item.get("cartons"), 1))) for item in items)
    total_cbm = sum(number(item.get("cartonLength")) * number(item.get("cartonWidth")) * number(item.get("cartonHeight")) * max(1, int(number(item.get("cartons"), 1))) / 1_000_000 for item in items)
    total_value = sum(number(item.get("quantity")) * number(item.get("unitPrice")) * max(1, int(number(item.get("cartons"), 1))) for item in items)
    full_address, full_address_inline = custom_full_address(shipment)
    values = dict(shipment)
    values.update({
        "fullAddress": full_address,
        "fullAddressInline": full_address_inline,
        "currency": next(iter(currencies)) if len(currencies) == 1 else ("MIXED" if currencies else safe_text(shipment.get("currency"), "USD")),
        "totalCartons": total_cartons, "totalQuantity": total_quantity,
        "totalGrossWeight": total_gross, "totalNetWeight": total_net,
        "totalCbm": total_cbm, "totalValue": total_value,
    })
    return values


def custom_fixed_output_value(record: dict, field: str, value):
    """Use the template's vocabulary for enums without treating examples as data."""
    if value in (None, ""):
        return None
    example = safe_text((record.get("fixedExamples") or {}).get(field))
    if field == "currency":
        currency = safe_text(value).upper()
        chinese_currency = {
            "USD": "美元", "EUR": "欧元", "GBP": "英镑", "CNY": "人民币",
            "JPY": "日元", "HKD": "港币", "CAD": "加元", "AUD": "澳元",
        }
        if example in chinese_currency.values() or any(word in example for word in chinese_currency.values()):
            return chinese_currency.get(currency, value)
    return value


def custom_item_rows(items: list, shipment: dict, row_mode: str) -> list[dict]:
    if row_mode == "carton":
        return expand_yuntuo_cartons(items, shipment)
    rows = []
    for item in items:
        row = dict(item)
        row["boxNumber"] = safe_text(row.get("boxNumber")) or safe_text(shipment.get("fbaNumber"))
        row["poNumber"] = safe_text(row.get("poNumber")) or safe_text(shipment.get("amazonReferenceId"))
        row["appointmentWindow"] = safe_text(row.get("appointmentWindow")) or safe_text(shipment.get("appointmentWindow"))
        rows.append(row)
    return rows


def custom_item_value(item: dict, field: str):
    cartons = max(1, int(number(item.get("cartons"), 1)))
    quantity = number(item.get("quantity"))
    if field == "totalQuantity":
        return quantity * cartons
    if field == "totalAmount":
        return quantity * cartons * number(item.get("unitPrice"))
    if field == "productWeightText":
        return safe_text(item.get("unitWeightText")) or product_weight_text(item)
    if field == "material":
        return joined_bilingual(item.get("materialZh"), item.get("materialEn"))
    if field == "purpose":
        return joined_bilingual(item.get("purposeZh"), item.get("purposeEn"))
    if field == "cbm":
        return number(item.get("cbm")) or number(item.get("cartonLength")) * number(item.get("cartonWidth")) * number(item.get("cartonHeight")) * cartons / 1_000_000
    if field == "image":
        return safe_text(item.get("image"))
    return item.get(field)


def missing_required_value(value) -> bool:
    return value is None or (isinstance(value, str) and not value.strip())


def custom_template_validation_errors(payload: dict, record: dict) -> list[str]:
    """Validate the required fields configured for this individual template."""
    mapping = record.get("mapping") or {}
    required = template_required_mapping(record)
    catalog = template_editor_field_catalog(record)
    labels = {
        category: {field["key"]: safe_text(field.get("label"), field["key"]) for field in fields}
        for category, fields in catalog.items()
    }
    errors = []
    fixed_mapping = mapping.get("fixed") or {}
    fixed_values = custom_fixed_values(payload)
    for field in required["fixed"]:
        if field in fixed_mapping and missing_required_value(fixed_values.get(field)):
            errors.append(f"必填字段未填写：{labels['fixed'].get(field, field)}")

    item_mapping = mapping.get("items") or {}
    mapped_columns = item_mapping.get("columns") or {}
    rows = custom_item_rows(payload.get("items", []), payload.get("shipment", {}), safe_text(item_mapping.get("rowMode"), "sku"))
    for field in required["items"]:
        if field not in mapped_columns:
            continue
        missing_rows = [str(index + 1) for index, row in enumerate(rows) if missing_required_value(custom_item_value(row, field))]
        if missing_rows:
            shown = "、".join(missing_rows[:8]) + ("等" if len(missing_rows) > 8 else "")
            errors.append(f"第 {shown} 行缺少必填字段：{labels['items'].get(field, field)}")
    return errors


def mapping_column(value) -> int | None:
    if isinstance(value, (int, float)) and int(value) > 0:
        return int(value)
    text = safe_text(value).upper()
    if text.isdigit() and int(text) > 0:
        return int(text)
    if re.fullmatch(r"[A-Z]{1,3}", text):
        return column_index_from_string(text)
    return None


def writable_template_cell(ws, row: int, column: int):
    """Return the writable anchor when a mapped target is inside a merged range."""
    cell = ws.cell(row, column)
    if not isinstance(cell, MergedCell):
        return cell
    for merged in ws.merged_cells.ranges:
        if merged.min_row <= row <= merged.max_row and merged.min_col <= column <= merged.max_col:
            return ws.cell(merged.min_row, merged.min_col)
    return cell


def expand_detail_formula(formula: str, detail_start: int, original_end: int, extra_rows: int) -> str:
    if extra_rows <= 0 or not isinstance(formula, str) or not formula.startswith("="):
        return formula
    pattern = re.compile(r"(\$?[A-Z]{1,3}\$?)(\d+):(\$?[A-Z]{1,3}\$?)(\d+)", re.IGNORECASE)
    def replace(match):
        start_row = int(match.group(2))
        end_row = int(match.group(4))
        if start_row >= detail_start and end_row == original_end:
            return f"{match.group(1)}{match.group(2)}:{match.group(3)}{end_row + extra_rows}"
        return match.group(0)
    return pattern.sub(replace, formula)


def export_custom_template(payload: dict, record: dict) -> tuple[bytes, str]:
    if not record.get("configured"):
        raise ValueError("该自定义模板尚未完成字段映射")
    template_path = custom_template_path(record)
    extension = template_path.suffix.casefold()
    # Load from memory so Windows never keeps the user's template file locked
    # after preview/export, which would otherwise make immediate deletion fail.
    workbook = load_workbook(io.BytesIO(template_path.read_bytes()), data_only=False, keep_vba=extension == ".xlsm")
    mapping = record.get("mapping", {})
    sheet_name = safe_text(mapping.get("sheet"))
    if sheet_name not in workbook.sheetnames:
        raise ValueError("映射的工作表不存在")
    ws = workbook[sheet_name]
    fixed_values = custom_fixed_values(payload)
    fixed_mapping = mapping.get("fixed") or {}
    # Some templates provide one large "完整收件地址" cell, while others only
    # name it "地址一" even though users expect the complete warehouse address.
    # Prefer the compact full address only in the latter case, so no address
    # component (especially the warehouse code) is silently discarded.
    if "address1" in fixed_mapping and "fullAddress" not in fixed_mapping:
        fixed_values["address1"] = fixed_values.get("fullAddressInline") or fixed_values.get("address1")
    pending_fixed_cells = {}
    fixed_priorities = {"fullAddress": 100, "address1": 90, "warehouseCode": 80}
    for order, (field, address) in enumerate(fixed_mapping.items()):
        coordinate = safe_text(address).upper()
        if re.fullmatch(r"[A-Z]{1,3}[1-9]\d*", coordinate):
            target = ws[coordinate]
            target = writable_template_cell(ws, target.row, target.column)
            value = custom_fixed_output_value(record, field, fixed_values.get(field))
            priority = fixed_priorities.get(field, 10)
            if value not in (None, ""):
                priority += 1000
            current = pending_fixed_cells.get(target.coordinate)
            if current is None or priority > current[0]:
                pending_fixed_cells[target.coordinate] = (priority, order, value)
    for coordinate, (_, _, value) in pending_fixed_cells.items():
        target = ws[coordinate]
        target.value = value if value not in (None, "") else None
        if isinstance(value, str) and "\n" in value:
            target.alignment = copy.copy(target.alignment)
            target.alignment = Alignment(
                horizontal=target.alignment.horizontal,
                vertical=target.alignment.vertical or "center",
                text_rotation=target.alignment.text_rotation,
                wrap_text=True,
                shrink_to_fit=target.alignment.shrink_to_fit,
                indent=target.alignment.indent,
            )

    item_mapping = mapping.get("items") or {}
    start_row = max(1, int(number(item_mapping.get("startRow"), 1)))
    reserved_rows = max(1, int(number(item_mapping.get("reservedRows"), 1)))
    row_mode = safe_text(item_mapping.get("rowMode"), "sku")
    rows = custom_item_rows(payload.get("items", []), payload.get("shipment", {}), row_mode)
    mapped_columns = {field: mapping_column(column) for field, column in (item_mapping.get("columns") or {}).items()}
    mapped_columns = {field: column for field, column in mapped_columns.items() if column is not None}
    if not mapped_columns:
        workbook.close()
        raise ValueError("请至少映射一个商品明细列")

    extra_rows = max(0, len(rows) - reserved_rows)
    source_formulas = {col: ws.cell(start_row, col).value for col in range(1, ws.max_column + 1) if isinstance(ws.cell(start_row, col).value, str) and ws.cell(start_row, col).value.startswith("=")}
    if extra_rows:
        ws.insert_rows(start_row + reserved_rows, amount=extra_rows)
        for target_row in range(start_row + reserved_rows, start_row + reserved_rows + extra_rows):
            copy_row_style(ws, start_row, target_row, 1, ws.max_column)
            for col, formula in source_formulas.items():
                try:
                    ws.cell(target_row, col).value = Translator(formula, origin=ws.cell(start_row, col).coordinate).translate_formula(ws.cell(target_row, col).coordinate)
                except Exception:
                    ws.cell(target_row, col).value = formula
        first_row_below_detail = start_row + reserved_rows + extra_rows
        for formula_row in range(first_row_below_detail, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(formula_row, col)
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    cell.value = expand_detail_formula(cell.value, start_row, start_row + reserved_rows - 1, extra_rows)

    image_column = mapped_columns.get("image")
    clear_count = max(reserved_rows, len(rows))
    # Some example templates contain broken placeholder pictures or icon
    # artifacts in price/link cells. Any picture anchored below the item header
    # in a mapped detail column is sample data, so remove it before adding the
    # real product images.
    mapped_detail_columns = set(mapped_columns.values())
    kept_images = []
    for drawing in getattr(ws, "_images", []):
        try:
            anchor = drawing.anchor
            row = anchor._from.row + 1 if not isinstance(anchor, str) else ws[anchor].row
            column = anchor._from.col + 1 if not isinstance(anchor, str) else ws[anchor].column
            if row >= start_row and column in mapped_detail_columns:
                continue
        except Exception:
            pass
        kept_images.append(drawing)
    ws._images = kept_images

    for row_index in range(start_row, start_row + clear_count):
        for field, column in mapped_columns.items():
            # Image placeholders are often formulas such as DISPIMG(...).
            # Clear the cell value for every mapped column; the real image is
            # inserted later as a drawing and does not need a formula/value.
            target = ws.cell(row_index, column)
            if not isinstance(target, MergedCell):
                target.value = None
    image_width = max(10, int(number(item_mapping.get("imageMaxWidth"), 72)))
    image_height = max(10, int(number(item_mapping.get("imageMaxHeight"), 50)))
    for index, item in enumerate(rows):
        target_row = start_row + index
        for field, column in mapped_columns.items():
            value = custom_item_value(item, field)
            if field == "image":
                add_cell_image(ws, f"{get_column_letter(column)}{target_row}", safe_text(value), image_width, image_height)
            else:
                target = writable_template_cell(ws, target_row, column)
                target.value = value if value not in (None, "") else None

    output = io.BytesIO()
    workbook.save(output)
    workbook.close()
    fba_number = safe_text(payload.get("shipment", {}).get("fbaNumber")) or "自定义发票"
    filename = f"{clean_filename(fba_number)}-{clean_filename(record.get('name'))}{extension}"
    return output.getvalue(), filename


def generate_workbook(payload: dict) -> tuple[bytes, str]:
    template_id = safe_text(payload.get("templateId"))
    shipment = payload.get("shipment", {})
    custom_record = custom_template_record(template_id)
    template_origin = safe_text(custom_record.get("migratedFromBuiltin")) if custom_record else template_id
    if template_origin in ("ruichi", "yuntuo"):
        if not safe_text(shipment.get("fbaNumber")):
            raise ValueError(f"{'锐驰' if template_origin == 'ruichi' else '云拓'}模板必须填写客户订单号(FBA#)")
        if not safe_text(shipment.get("amazonReferenceId")):
            message = "（将自动作为 PO Number）" if template_origin == "yuntuo" else ""
            raise ValueError(f"{'锐驰' if template_origin == 'ruichi' else '云拓'}模板必须填写 Amazon Reference ID{message}")
    if template_id == "ruichi":
        return export_ruichi(payload)
    if template_id == "yuntuo":
        if not safe_text(shipment.get("service")):
            raise ValueError("云拓模板必须选择服务渠道")
        missing_export_sku = [(safe_text(item.get("sku")) or f"第 {index + 1} 个商品") for index, item in enumerate(payload.get("items", [])) if not safe_text(item.get("exportSku"))]
        if missing_export_sku:
            raise ValueError(f"以下商品缺少导出 SKU：{', '.join(missing_export_sku)}")
        return export_yuntuo(payload)
    if custom_record is not None:
        if template_origin == "yuntuo":
            if not safe_text(shipment.get("service")):
                raise ValueError("云拓模板必须选择服务渠道")
            missing_export_sku = [(safe_text(item.get("sku")) or f"第 {index + 1} 个商品") for index, item in enumerate(payload.get("items", [])) if not safe_text(item.get("exportSku"))]
            if missing_export_sku:
                raise ValueError(f"以下商品缺少导出 SKU：{', '.join(missing_export_sku)}")
        required_errors = custom_template_validation_errors(payload, custom_record)
        if required_errors:
            raise ValueError("；".join(required_errors))
        return export_custom_template(payload, custom_record)
    raise ValueError("未知模板")


def excel_color(color) -> str | None:
    if color is None:
        return None
    value = getattr(color, "rgb", None)
    if isinstance(value, str) and len(value) >= 6:
        return f"#{value[-6:]}"
    indexed = getattr(color, "indexed", None)
    fallback = {
        0: "#000000", 1: "#ffffff", 2: "#ff0000", 3: "#00ff00",
        4: "#0000ff", 5: "#ffff00", 6: "#ff00ff", 7: "#00ffff",
        8: "#000000", 9: "#ffffff", 10: "#ff0000", 11: "#00ff00",
        12: "#0000ff", 13: "#ffff00", 14: "#ff00ff", 15: "#00ffff",
    }
    return fallback.get(indexed)


def border_css(side, css_name: str) -> str | None:
    if side is None or side.style is None:
        return None
    width = "2px" if side.style in ("medium", "thick", "double") else "1px"
    line = "dashed" if "dash" in side.style else "solid"
    color = excel_color(side.color) or "#c9d1dc"
    return f"border-{css_name}:{width} {line} {color}"


def cell_css(cell) -> str:
    rules = []
    fill = excel_color(cell.fill.fgColor) if cell.fill and cell.fill.fill_type else None
    if fill:
        rules.append(f"background-color:{fill}")
    if cell.font:
        if cell.font.bold:
            rules.append("font-weight:700")
        if cell.font.italic:
            rules.append("font-style:italic")
        if cell.font.sz:
            rules.append(f"font-size:{max(8, min(float(cell.font.sz), 22))}px")
        color = excel_color(cell.font.color)
        if color:
            rules.append(f"color:{color}")
    if cell.alignment:
        if cell.alignment.horizontal:
            horizontal = {"centerContinuous": "center", "distributed": "center", "fill": "left"}.get(
                cell.alignment.horizontal, cell.alignment.horizontal
            )
            rules.append(f"text-align:{horizontal}")
        if cell.alignment.vertical:
            rules.append(f"vertical-align:{cell.alignment.vertical}")
        if cell.alignment.wrap_text:
            rules.append("white-space:pre-wrap")
    if cell.border:
        for value in (
            border_css(cell.border.left, "left"), border_css(cell.border.right, "right"),
            border_css(cell.border.top, "top"), border_css(cell.border.bottom, "bottom"),
        ):
            if value:
                rules.append(value)
    return ";".join(rules)


def formula_display(ws, formula: str):
    sum_match = re.fullmatch(r"=SUM\(([A-Z]+)(\d+):([A-Z]+)(\d+)\)", formula, re.IGNORECASE)
    if sum_match:
        start_col, start_row, end_col, end_row = sum_match.groups()
        total = 0.0
        for row in ws[f"{start_col}{start_row}":f"{end_col}{end_row}"]:
            for cell in row:
                if isinstance(cell.value, (int, float)):
                    total += cell.value
        return total
    multiply_match = re.fullmatch(r"=([A-Z]+\d+)\*([A-Z]+\d+)", formula, re.IGNORECASE)
    if multiply_match:
        return number(ws[multiply_match.group(1)].value) * number(ws[multiply_match.group(2)].value)
    return formula


def workbook_preview(content: bytes, filename: str) -> dict:
    workbook = load_workbook(io.BytesIO(content), data_only=False)
    sheets = []
    for ws in workbook.worksheets:
        max_row = min(max(ws.max_row, 1), 300)
        max_col = min(max(ws.max_column, 1), 60)
        merges = []
        for merged in ws.merged_cells.ranges:
            merges.append({
                "minRow": merged.min_row, "maxRow": merged.max_row,
                "minCol": merged.min_col, "maxCol": merged.max_col,
            })
        cells = []
        for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
            for cell in row:
                value = cell.value
                display = formula_display(ws, value) if isinstance(value, str) and value.startswith("=") else value
                cells.append({
                    "row": cell.row, "col": cell.column, "address": cell.coordinate,
                    "value": "" if value is None else value,
                    "display": "" if display is None else display,
                    "formula": isinstance(value, str) and value.startswith("="),
                    "style": cell_css(cell),
                })
        images = []
        for drawing in getattr(ws, "_images", []):
            try:
                anchor = drawing.anchor
                if isinstance(anchor, str):
                    match = re.match(r"([A-Z]+)(\d+)", anchor)
                    if not match:
                        continue
                    col = ws[anchor].column
                    row = ws[anchor].row
                else:
                    col = anchor._from.col + 1
                    row = anchor._from.row + 1
                raw = drawing._data()
                extension = (drawing.format or "png").lower()
                mime = "jpeg" if extension in ("jpg", "jpeg") else extension
                images.append({
                    "row": row, "col": col,
                    "src": f"data:image/{mime};base64,{base64.b64encode(raw).decode('ascii')}",
                })
            except Exception:
                continue
        sheets.append({
            "name": ws.title, "maxRow": max_row, "maxCol": max_col,
            "merges": merges, "cells": cells, "images": images,
            "columnWidths": {
                str(col): max(35, min(280, int((ws.column_dimensions[get_column_letter(col)].width or 10) * 7)))
                for col in range(1, max_col + 1)
            },
            "rowHeights": {
                str(row): max(20, min(240, int((ws.row_dimensions[row].height or 15) * 1.333)))
                for row in range(1, max_row + 1)
            },
            "hiddenRows": [row for row in range(1, max_row + 1) if ws.row_dimensions[row].hidden],
            "hiddenCols": [col for col in range(1, max_col + 1) if ws.column_dimensions[get_column_letter(col)].hidden],
        })
    workbook.close()
    return {"filename": filename, "sheets": sheets}


def coerce_edited_value(value):
    if not isinstance(value, str):
        return value
    stripped = value.strip()
    if not stripped:
        return None
    if stripped.startswith("="):
        return stripped
    if re.fullmatch(r"-?(0|[1-9]\d*)(\.\d+)?", stripped):
        if stripped.startswith("0") and len(stripped) > 1 and not stripped.startswith("0."):
            return value
        return float(stripped) if "." in stripped else int(stripped)
    return value


def apply_cell_edits(content: bytes, edits: list[dict]) -> bytes:
    if not edits:
        return content
    workbook = load_workbook(io.BytesIO(content), data_only=False)
    for edit in edits:
        sheet_name = safe_text(edit.get("sheet"))
        address = safe_text(edit.get("address")).upper()
        if sheet_name not in workbook.sheetnames or not re.fullmatch(r"[A-Z]{1,3}[1-9]\d{0,5}", address):
            continue
        workbook[sheet_name][address] = coerce_edited_value(edit.get("value"))
    output = io.BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def channels() -> list[dict]:
    file_path = TEMPLATES["yuntuo"]["file"]
    try:
        workbook = load_workbook(file_path, read_only=True, data_only=True)
        ws = workbook["渠道名称与代码"]
        values = []
        for row in range(2, ws.max_row + 1):
            name = safe_text(ws.cell(row, 1).value)
            code = safe_text(ws.cell(row, 2).value)
            if name and code:
                label = name if code.casefold() in name.casefold() else f"{name}（{code}）"
                values.append({"name": name, "code": code, "label": label})
        workbook.close()
        return values
    except Exception:
        return []


class InvoiceHandler(SimpleHTTPRequestHandler):
    server_version = f"FreightInvoice/{APP_VERSION}"

    def log_message(self, format, *args):
        if sys.stdout is not None:
            sys.stdout.write("[%s] %s\n" % (self.log_date_time_string(), format % args))

    def send_json(self, payload: dict, status=HTTPStatus.OK):
        body = json.dumps(payload, ensure_ascii=False).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def do_GET(self):
        parsed_url = urlparse(self.path)
        route = unquote(parsed_url.path)
        if route == "/api/backup":
            content, filename, manifest = create_full_backup()
            self.send_response(HTTPStatus.OK)
            self.send_header("Content-Type", "application/zip")
            self.send_header("Content-Disposition", 'attachment; filename="freight-invoice-backup.zip"')
            self.send_header("X-Filename", base64.b64encode(filename.encode("utf-8")).decode("ascii"))
            self.send_header("X-Backup-Counts", base64.b64encode(json.dumps(manifest["counts"], ensure_ascii=False).encode("utf-8")).decode("ascii"))
            self.send_header("Content-Length", str(len(content)))
            self.end_headers()
            self.wfile.write(content)
            return
        if route == "/api/products":
            self.send_json({"products": read_products()})
            return
        if route == "/api/field-catalog":
            self.send_json(field_catalog(include_disabled=True))
            return
        if route == "/api/config":
            self.send_json({
                "templates": template_catalog(include_pending=False),
                "channels": channels(),
                "warehouseCount": len(read_warehouses()),
                "appVersion": APP_VERSION,
                "appEdition": APP_EDITION,
                "storageMode": "服务端共享数据" if self.server.server_address[0] not in ("127.0.0.1", "localhost") else "本机数据",
            })
            return
        if route == "/api/templates":
            catalog = field_catalog()
            self.send_json({
                "templates": template_catalog(include_pending=True),
                "fixedFields": catalog["fixed"],
                "itemFields": catalog["items"],
            })
            return
        if route.startswith("/api/templates/"):
            template_suffix = route.removeprefix("/api/templates/")
            wants_preview = template_suffix.endswith("/preview")
            template_id = template_suffix.removesuffix("/preview") if wants_preview else template_suffix
            record = custom_template_record(template_id)
            if record is None:
                self.send_json({"error": "自定义模板不存在"}, HTTPStatus.NOT_FOUND)
            elif wants_preview:
                content = custom_template_path(record).read_bytes()
                self.send_json(workbook_preview(content, safe_text(record.get("originalFilename"), "template.xlsx")))
            else:
                catalog = template_editor_field_catalog(record)
                self.send_json({
                    "template": custom_template_detail(record),
                    "fixedFields": catalog["fixed"],
                    "itemFields": catalog["items"],
                })
            return
        if route == "/api/warehouses":
            params = parse_qs(parsed_url.query)
            query = safe_text(params.get("q", [""])[0])
            try:
                limit = int(params.get("limit", ["30"])[0])
            except (TypeError, ValueError):
                limit = 30
            warehouses, total = search_warehouses(query, limit)
            self.send_json({"warehouses": warehouses, "total": total})
            return
        if route == "/api/history":
            self.send_json({"history": [export_history_summary(record) for record in read_export_history()]})
            return
        if route == "/api/draft":
            self.send_json({"draft": read_invoice_draft()})
            return
        if route.startswith("/api/history/"):
            history_suffix = route.removeprefix("/api/history/")
            wants_file = history_suffix.endswith("/file")
            record_id = history_suffix.removesuffix("/file") if wants_file else history_suffix
            record = next((item for item in read_export_history() if safe_text(item.get("id")) == record_id), None)
            if record is None:
                self.send_json({"error": "历史记录不存在"}, HTTPStatus.NOT_FOUND)
            elif wants_file:
                stored_file = history_file_path(record_id)
                if not stored_file.is_file():
                    self.send_json({"error": "该历史记录没有保存 Excel 文件"}, HTTPStatus.NOT_FOUND)
                    return
                content = stored_file.read_bytes()
                filename = safe_text(record.get("filename"), "历史发票.xlsx")
                self.send_response(HTTPStatus.OK)
                self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                self.send_header("Content-Disposition", 'attachment; filename="invoice.xlsx"')
                self.send_header("X-Filename", base64.b64encode(filename.encode("utf-8")).decode("ascii"))
                self.send_header("Content-Length", str(len(content)))
                self.end_headers()
                self.wfile.write(content)
            else:
                self.send_json({"record": record})
            return
        if route == "/health":
            self.send_json({"status": "ok", "app": "freight-invoice-studio", "version": APP_VERSION, "edition": APP_EDITION, "time": datetime.now().isoformat()})
            return
        if route == "/":
            route = "/index.html"
        target = (STATIC_DIR / route.lstrip("/")).resolve()
        if STATIC_DIR.resolve() not in target.parents and target != STATIC_DIR.resolve():
            self.send_error(HTTPStatus.FORBIDDEN)
            return
        if not target.is_file():
            self.send_error(HTTPStatus.NOT_FOUND)
            return
        content = target.read_bytes()
        content_type = mimetypes.guess_type(target.name)[0] or "application/octet-stream"
        if content_type.startswith("text/") or content_type in ("application/javascript", "application/json"):
            content_type += "; charset=utf-8"
        self.send_response(HTTPStatus.OK)
        self.send_header("Content-Type", content_type)
        self.send_header("Content-Length", str(len(content)))
        self.end_headers()
        self.wfile.write(content)

    def do_POST(self):
        route = unquote(urlparse(self.path).path)
        if route == "/api/restore":
            try:
                content_length = int(self.headers.get("Content-Length", "0"))
                if content_length <= 0 or content_length > BACKUP_MAX_BYTES:
                    raise ValueError("备份文件为空或超过 1GB")
                result = restore_full_backup(self.rfile.read(content_length))
                self.send_json({"status": "ok", "restored": result})
            except Exception as exc:
                self.send_json({"error": str(exc)}, HTTPStatus.BAD_REQUEST)
            return
        is_template_upload = route == "/api/templates/upload"
        is_field_catalog_update = route == "/api/field-catalog"
        mapping_match = re.fullmatch(r"/api/templates/(tpl-[0-9a-f]{10})/mapping", route)
        reanalyze_match = re.fullmatch(r"/api/templates/(tpl-[0-9a-f]{10})/reanalyze", route)
        convert_match = re.fullmatch(r"/api/templates/(ruichi|yuntuo)/convert", route)
        if route not in ("/api/export", "/api/preview", "/api/products", "/api/draft", "/api/warehouses", "/api/products/bulk-delete", "/api/templates/bulk-delete", "/api/history/bulk-delete", "/api/data/clear-all") and not is_template_upload and not is_field_catalog_update and not mapping_match and not reanalyze_match and not convert_match:
            self.send_error(HTTPStatus.NOT_FOUND)
            return
        try:
            content_length = int(self.headers.get("Content-Length", "0"))
            if content_length <= 0 or content_length > 35 * 1024 * 1024:
                raise ValueError("请求数据为空或过大")
            payload = json.loads(self.rfile.read(content_length).decode("utf-8"))
            if route == "/api/data/clear-all":
                if safe_text(payload.get("confirmation")) != "CLEAR_ALL_DATA":
                    raise ValueError("清空确认信息不正确")
                cleared = clear_all_user_data()
                self.send_json({"status": "ok", "cleared": cleared})
                return
            if route == "/api/warehouses":
                warehouse = upsert_warehouse(payload)
                self.send_json({"status": "ok", "warehouse": warehouse, "total": len(read_warehouses())})
                return
            if is_field_catalog_update:
                field = upsert_field_definition(payload)
                self.send_json({"status": "ok", "field": field})
                return
            if is_template_upload:
                encoded = safe_text(payload.get("data"))
                if "," in encoded:
                    encoded = encoded.split(",", 1)[1]
                try:
                    content = base64.b64decode(encoded, validate=True)
                except (ValueError, base64.binascii.Error) as exc:
                    raise ValueError("模板文件数据无效") from exc
                record = create_custom_template(payload.get("name"), safe_text(payload.get("filename"), "template.xlsx"), content)
                self.send_json({"status": "ok", "template": record})
                return
            if mapping_match:
                record = update_custom_template_mapping(mapping_match.group(1), payload)
                self.send_json({"status": "ok", "template": record})
                return
            if reanalyze_match:
                record = reanalyze_custom_template(reanalyze_match.group(1))
                self.send_json({"status": "ok", "template": record})
                return
            if convert_match:
                record = convert_builtin_to_custom(convert_match.group(1))
                self.send_json({"status": "ok", "template": record})
                return
            if route == "/api/products/bulk-delete":
                count = delete_products(payload.get("ids") if isinstance(payload.get("ids"), list) else [])
                self.send_json({"status": "ok", "deleted": count})
                return
            if route == "/api/templates/bulk-delete":
                count = delete_custom_templates(payload.get("ids") if isinstance(payload.get("ids"), list) else [])
                self.send_json({"status": "ok", "deleted": count})
                return
            if route == "/api/history/bulk-delete":
                count = delete_export_histories(payload.get("ids") if isinstance(payload.get("ids"), list) else [])
                self.send_json({"status": "ok", "deleted": count})
                return
            if route == "/api/products":
                write_products(payload.get("products", []))
                self.send_json({"status": "ok", "count": len(payload.get("products", []))})
                return
            if route == "/api/draft":
                result = write_invoice_draft(payload)
                self.send_json({"status": "ok", **result})
                return
            content, filename = generate_workbook(payload)
            if route == "/api/preview":
                self.send_json(workbook_preview(content, filename))
                return
            content = apply_cell_edits(content, payload.get("cellEdits", []))
            history_record = save_export_history(payload, filename, content)
            encoded_name = filename.encode("utf-8").hex()
            self.send_response(HTTPStatus.OK)
            self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            self.send_header("Content-Disposition", f'attachment; filename="invoice.xlsx"; filename-hex="{encoded_name}"')
            self.send_header("X-Filename", base64.b64encode(filename.encode("utf-8")).decode("ascii"))
            self.send_header("X-History-Id", history_record["id"])
            self.send_header("Content-Length", str(len(content)))
            self.end_headers()
            self.wfile.write(content)
        except Exception as exc:
            self.send_json({"error": str(exc)}, HTTPStatus.BAD_REQUEST)

    def do_DELETE(self):
        route = unquote(urlparse(self.path).path)
        if route == "/api/draft":
            clear_invoice_draft()
            self.send_json({"status": "ok"})
            return
        field_match = re.fullmatch(r"/api/field-catalog/(fixed|items)/([A-Za-z][A-Za-z0-9_]*)", route)
        if field_match:
            if delete_field_definition(field_match.group(1), field_match.group(2)):
                self.send_json({"status": "ok"})
            else:
                self.send_json({"error": "字段不存在"}, HTTPStatus.NOT_FOUND)
            return
        warehouse_match = re.fullmatch(r"/api/warehouses/(wh-[0-9a-f]{12})", route)
        if warehouse_match:
            if delete_warehouse(warehouse_match.group(1)):
                self.send_json({"status": "ok", "total": len(read_warehouses())})
            else:
                self.send_json({"error": "仓库不存在"}, HTTPStatus.NOT_FOUND)
            return
        template_match = re.fullmatch(r"/api/templates/(tpl-[0-9a-f]{10}|ruichi|yuntuo)", route)
        if template_match:
            template_id = template_match.group(1)
            removed = disable_builtin_template(template_id) if template_id in TEMPLATES else delete_custom_template(template_id)
            if removed:
                self.send_json({"status": "ok"})
            else:
                self.send_json({"error": "自定义模板不存在"}, HTTPStatus.NOT_FOUND)
            return
        if not route.startswith("/api/history/"):
            self.send_error(HTTPStatus.NOT_FOUND)
            return
        record_id = route.removeprefix("/api/history/")
        if delete_export_history(record_id):
            self.send_json({"status": "ok"})
        else:
            self.send_json({"error": "历史记录不存在"}, HTTPStatus.NOT_FOUND)


def main():
    args = sys.argv[1:]
    requested_port = None
    requested_host = os.environ.get("FREIGHT_INVOICE_HOST", "127.0.0.1")
    if args and args[0].isdigit():
        requested_port = int(args[0])
    if "--port" in args and args.index("--port") + 1 < len(args):
        requested_port = int(args[args.index("--port") + 1])
    if "--host" in args and args.index("--host") + 1 < len(args):
        requested_host = args[args.index("--host") + 1]
    if getattr(sys, "frozen", False) and requested_port is None and requested_host in ("127.0.0.1", "localhost"):
        opener = build_opener(ProxyHandler({}))
        for candidate in range(8765, 8776):
            try:
                with opener.open(f"http://127.0.0.1:{candidate}/health", timeout=0.2) as response:
                    status = json.loads(response.read().decode("utf-8"))
                if status.get("status") == "ok" and status.get("app") == "freight-invoice-studio":
                    webbrowser.open(f"http://127.0.0.1:{candidate}")
                    return
            except Exception:
                continue
    EXPORT_DIR.mkdir(exist_ok=True)
    DATA_DIR.mkdir(exist_ok=True)
    candidates = [requested_port] if requested_port is not None else list(range(8765, 8776))
    server = None
    port = candidates[0]
    for candidate in candidates:
        try:
            server = ThreadingHTTPServer((requested_host, candidate), InvoiceHandler)
            port = candidate
            break
        except OSError:
            continue
    if server is None:
        raise RuntimeError(f"无法启动服务：{requested_host} 的候选端口均不可用")
    public_host = "127.0.0.1" if requested_host in ("0.0.0.0", "::") else requested_host
    if sys.stdout is not None:
        print(f"货代发票工作台已启动：http://{public_host}:{port}")
    if getattr(sys, "frozen", False) and requested_host in ("127.0.0.1", "localhost") and os.environ.get("FREIGHT_INVOICE_NO_BROWSER") != "1":
        threading.Timer(0.8, lambda: webbrowser.open(f"http://{public_host}:{port}")).start()
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        pass
    finally:
        server.server_close()


if __name__ == "__main__":
    main()
